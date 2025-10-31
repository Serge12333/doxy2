import tkinter
from tkinter import *
from tkinter import messagebox
from tkinter.ttk import *
from tkinter import ttk
import tkinter as tk
import os
from mailmerge import MailMerge
from transliterate import translit
import tkentrycomplete
import json
import sys
from datetime import datetime, timedelta, date
import shutil
import openpyxl
from openpyxl.styles import numbers
from tkinter import simpledialog
from decimal import Decimal
import uuid
from decimal import Decimal, getcontext
from num2words import num2words
import numpy as np
from collections import defaultdict


# --- Globals ---
# Main container for dynamically created widgets
dynamic_frame = None
# Global dictionary to hold the state of dynamically created checkboxes
checkbox_vars = {}
# Global dictionary to hold selections from main-key comboboxes
main_key_selections = {}
getcontext().prec = 28

ALL_TAG_VALUES = {}   # current values of all interactive tags (поля/списки/чекбоксы/комбобоксы)
NUMBER_LABELS = {}


class LocalizedAskString(simpledialog._QueryString):
    def body(self, master):
        self.result = None
        return super().body(master)

    def buttonbox(self):
        box = tk.Frame(self)

        w = tk.Button(box, text="ОК", width=10, command=self.ok, default=tk.ACTIVE)
        w.pack(side=tk.LEFT, padx=5, pady=5)
        w = tk.Button(box, text="Отмена", width=10, command=self.cancel)
        w.pack(side=tk.LEFT, padx=5, pady=5)

        self.bind("<Return>", self.ok)
        self.bind("<Escape>", self.cancel)

        box.pack()

def askstring_localized(title, prompt, **kwargs):
    d = LocalizedAskString(title, prompt, **kwargs)
    return d.result


# Determine the base directory
# Determine the base directory
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)  # Dir containing the .exe
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))  # Dir containing the script

PROJECTS_FILE = os.path.join(BASE_DIR, "projects.json")
CURRENT_PROJECT = None


def load_projects():
    if not os.path.exists(PROJECTS_FILE):
        return {"projects": [], "last_opened": None}
    with open(PROJECTS_FILE, "r", encoding="utf8") as f:
        return json.load(f)


def save_projects(data):
    with open(PROJECTS_FILE, "w", encoding="utf8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)


def create_project(name):
    """Создаёт новый проект: json конфиги + папку documents/templates/<project_name> + documents/processed/<project_name>."""
    projects = load_projects()
    if any(p["name"] == name for p in projects["projects"]):
        messagebox.showerror("Ошибка", "Проект с таким именем уже существует.")
        return False

    # --- JSON config folder ---
    project_json_dir = os.path.join(BASE_DIR, "json", name)
    os.makedirs(project_json_dir, exist_ok=True)

    # Create empty config files
    for fname in [
        "fields_config.json",
        "combobox_regular.json",
        "combobox_mainkey.json",
        "combination_config.json",
        "rules_config.json",
        "all_tags.json",
    ]:
        with open(os.path.join(project_json_dir, fname), "w", encoding="utf-8") as f:
            json.dump([], f, ensure_ascii=False, indent=4)

    # --- Templates folder for this project ---
    project_templates_dir = os.path.join(BASE_DIR, "documents", "templates", name)
    os.makedirs(project_templates_dir, exist_ok=True)

    # Copy base templates from documents/templates/_base
    base_templates_dir = os.path.join(BASE_DIR, "documents", "templates", "_base")
    if os.path.exists(base_templates_dir):
        for file in os.listdir(base_templates_dir):
            src = os.path.join(base_templates_dir, file)
            dst = os.path.join(project_templates_dir, file)
            if os.path.isfile(src):
                shutil.copy2(src, dst)

    # --- Processed folder ---
    project_processed_dir = os.path.join(BASE_DIR, "documents", "processed", name)
    os.makedirs(project_processed_dir, exist_ok=True)

    # --- Update projects.json ---
    projects["projects"].append({"name": name, "autoload": False})
    save_projects(projects)
    return True



def delete_project(name):
    projects = load_projects()
    projects["projects"] = [p for p in projects["projects"] if p["name"] != name]
    if projects.get("last_opened") == name:
        projects["last_opened"] = None
    save_projects(projects)

    # Delete only JSONs
    json_dir = os.path.join(BASE_DIR, "json", name)
    if os.path.exists(json_dir):
        shutil.rmtree(json_dir)


def set_autoload(name, enabled):
    projects = load_projects()
    for p in projects["projects"]:
        p["autoload"] = (p["name"] == name) if enabled else False
    save_projects(projects)


def get_autoload_project():
    projects = load_projects()
    for p in projects["projects"]:
        if p.get("autoload", False):
            return p["name"]
    return None



# The JSON folder is inside the base directory
def set_current_project(name):
    """Активирует проект и перенастраивает все пути."""
    global CURRENT_PROJECT, JSON_DIR, TEMPLATES_DIR, PROCESSED_DIR
    global FIELDS_CONFIG_PATH, COMBOBOX_REGULAR_PATH, COMBOBOX_MAINKEY_PATH
    global COMBINATION_CONFIG_PATH, RULES_CONFIG_PATH, ALL_TAGS_OUTPUT_PATH
    global NUMBER_CONFIG_PATH, INPUT_STATE_PATH

    # Remove the `if` statement that calls `prompt_before_switch`
    # and all the code inside of it.

    CURRENT_PROJECT = name

    JSON_DIR = os.path.join(BASE_DIR, "json", CURRENT_PROJECT)
    TEMPLATES_DIR = os.path.join(BASE_DIR, "documents", "templates", CURRENT_PROJECT)
    PROCESSED_DIR = os.path.join(BASE_DIR, "documents", "processed", CURRENT_PROJECT)

    # Make sure dirs exist
    os.makedirs(JSON_DIR, exist_ok=True)
    os.makedirs(TEMPLATES_DIR, exist_ok=True)
    os.makedirs(PROCESSED_DIR, exist_ok=True)

    FIELDS_CONFIG_PATH = os.path.join(JSON_DIR, "fields_config.json")
    COMBOBOX_REGULAR_PATH = os.path.join(JSON_DIR, "combobox_regular.json")
    COMBOBOX_MAINKEY_PATH = os.path.join(JSON_DIR, "combobox_mainkey.json")
    COMBINATION_CONFIG_PATH = os.path.join(JSON_DIR, "combination_config.json")
    RULES_CONFIG_PATH = os.path.join(JSON_DIR, "rules_config.json")
    ALL_TAGS_OUTPUT_PATH = os.path.join(JSON_DIR, "all_tags.json")
    NUMBER_CONFIG_PATH = os.path.join(JSON_DIR, "number_config.json")
    INPUT_STATE_PATH = os.path.join(JSON_DIR, "input_state.json")


IMPORT_FLD = os.path.join(BASE_DIR, 'import_fld')

# --- Utility and Core Logic Functions ---



def _onKeyRelease(event):
    """Handles Ctrl+C, Ctrl+V, Ctrl+X for entry widgets."""
    ctrl = (event.state & 0x4) != 0
    if event.keycode == 88 and ctrl and event.keysym.lower() != "x":
        event.widget.event_generate("<<Cut>>")
    elif event.keycode == 86 and ctrl and event.keysym.lower() != "v":
        event.widget.event_generate("<<Paste>>")
    elif event.keycode == 67 and ctrl and event.keysym.lower() != "c":
        event.widget.event_generate("<<Copy>>")


def load_json(file_path, dict_name):
    """Loads data from a JSON file, creating it if it doesn't exist."""
    try:
        with open(file_path, 'r', encoding='utf8') as f:
            return json.load(f)
    except FileNotFoundError:
        with open(file_path, 'w', encoding='utf8') as f:
            json.dump([], f, ensure_ascii=False, indent=4)
        messagebox.showwarning("Информация", f"Создан новый файл конфигурации: {file_path}")
        return []
    except json.JSONDecodeError:
        messagebox.showwarning("Предупреждение", f"Файл {file_path} поврежден. Инициализация пустой конфигурации.")
        with open(file_path, 'w', encoding='utf8') as f:
            json.dump([], f, ensure_ascii=False, indent=4)
        return []


def save_json(file_path, data):
    """Safely writes data to a JSON file to prevent corruption."""
    temp_file = file_path + '.tmp'
    try:
        with open(temp_file, 'w', encoding='utf8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        os.replace(temp_file, file_path)
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось сохранить конфигурацию в {file_path}: {e}")
        if os.path.exists(temp_file):
            os.remove(temp_file)


def save_input_state():
    """Save current user input values to input_state.json."""
    if not dynamic_frame or not dynamic_frame.winfo_exists():
        return  # Nothing to save
    state = get_current_widget_values(dynamic_frame)
    if isinstance(state, dict):
        save_json(INPUT_STATE_PATH, state)


def load_input_state():
    """Load previously saved input values (if any)."""
    try:
        data = load_json(INPUT_STATE_PATH, 'input_state')
        if isinstance(data, dict):
            return data
        else:
            return {}
    except Exception:
        return {}

def _norm_in_to_decimal_str(v: object) -> str:
    """
    Normalize any user/input value to a Decimal-compatible string with dot as separator.
    Accepts '12,3' or '12.3'; non-numeric -> '0'.
    """
    s = str(v).strip().replace("\u00A0", "").replace(" ", "")
    if not s:
        return "0"
    s = s.replace(",", ".")
    try:
        # Validation only; return the normalized string if Decimal accepts it
        Decimal(s)
        return s
    except Exception:
        return "0"

def _fmt_out_from_decimal(d: Decimal) -> str:
    """
    Always show two decimals with a comma (e.g., 12,00).
    """
    try:
        q = d.quantize(Decimal("0.01"))
        return str(q).replace(".", ",")
    except Exception:
        return "0,00"

def update_number_labels():
    """
    Recalculate ALL Число tags and update their labels.
    Supports nested numbers (a number using another number) via a few resolving passes.
    """
    try:
        numbers = load_number_config()
    except Exception:
        numbers = []

    # quick lookup for names
    number_names = {n.get("name") for n in numbers if "name" in n}
    results = {name: "0,00" for name in number_names}

    # Do a few passes to resolve dependencies between numbers
    for _ in range(5):
        for num in numbers:
            name = num.get("name")
            seq = num.get("sequence", [])
            parts = []

            for elem in seq:
                # operators / brackets
                if elem in ["+", "-", "*", "/", "(", ")"]:
                    parts.append(elem)
                    continue

                # numeric literal
                if isinstance(elem, (int, float)) or str(elem).replace(",", "").replace(".", "").lstrip("-").isdigit():
                    s = _norm_in_to_decimal_str(elem)
                    parts.append(f"Decimal('{s}')")
                    continue

                # reference to another tag (entry/combobox/checkbox)
                if elem in ALL_TAG_VALUES:
                    s = _norm_in_to_decimal_str(ALL_TAG_VALUES.get(elem, "0"))
                    parts.append(f"Decimal('{s}')")
                    continue

                # reference to a subkey inside main_key selections
                found_subkey = False
                for subdict in main_key_selections.values():
                    if elem in subdict:
                        s = _norm_in_to_decimal_str(subdict.get(elem, "0"))
                        parts.append(f"Decimal('{s}')")
                        found_subkey = True
                        break
                if found_subkey:
                    continue

                # reference to another number
                if elem in number_names:
                    s = _norm_in_to_decimal_str(results.get(elem, "0,00"))
                    parts.append(f"Decimal('{s}')")
                    continue

                # anything else -> 0
                parts.append("Decimal('0')")

            expr = " ".join(parts) if parts else "Decimal('0')"
            try:
                val = eval(expr, {"Decimal": Decimal})
                results[name] = _fmt_out_from_decimal(val)
            except Exception:
                results[name] = "0,00"

    # Push results to visible labels
    for name, value in results.items():
        lbl = NUMBER_LABELS.get(name)
        if lbl:
            lbl.config(text=value)

def load_number_config():
    """Loads Число tags configuration from JSON."""
    return load_json(NUMBER_CONFIG_PATH, 'number_config')

def save_number_config(data):
    """Saves Число tags configuration to JSON."""
    save_json(NUMBER_CONFIG_PATH, data)

def populate_tags_listbox(tags_listbox):
    """Populates the tags listbox in the constructor window."""
    for i in tags_listbox.get_children():
        tags_listbox.delete(i)

    all_items = []
    all_items.extend(
        [(f['name'], f['type'], f.get('tag_type', 'поле')) for f in load_json(FIELDS_CONFIG_PATH, 'fields_config')])
    all_items.extend([(c['name'], c['type'], c.get('tag_type', 'комбобокс')) for c in
                      load_json(COMBOBOX_REGULAR_PATH, 'combobox_regular')])
    all_items.extend([(c['name'], c['type'], c.get('tag_type', 'список')) for c in
                      load_json(COMBOBOX_MAINKEY_PATH, 'combobox_mainkey')])

    try:
        combination_config = load_combination_config()
        for combo in combination_config:
            all_items.append((combo['name'], combo['type'], combo.get('tag_type', 'сочетание')))
    except Exception as e:
        messagebox.showerror("Ошибка", f"Ошибка при загрузке combination_config: {str(e)}")

    # Load numbers
    try:
        number_config = load_number_config()
        for num in number_config:
            all_items.append((num['name'], num['type'], num.get('tag_type', 'число')))
    except Exception as e:
        messagebox.showerror("Ошибка", f"Ошибка при загрузке number_config: {str(e)}")

    all_items.sort(key=lambda x: x[0])
    for item in all_items:
        tags_listbox.insert("", "end", values=item)


def get_current_widget_values(frame):
    """Saves the current values from widgets in the given frame."""
    state = {}
    if not frame:
        return state
    for widget in frame.winfo_children():
        if not hasattr(widget, '_name'):
            continue
        widget_name = widget._name
        if isinstance(widget, (tk.Entry, tkentrycomplete.Combobox)):
            state[widget_name] = widget.get()
        elif isinstance(widget, ttk.Checkbutton):
            var = checkbox_vars.get(widget_name)
            if var:
                state[widget_name] = var.get()  # Save the integer value (0 or 1)
    return state


def refresh_all_windows(listbox_to_refresh):
    """Refreshes the dynamic widgets in the main window and the constructor listbox."""
    global dynamic_frame

    # 1. Save current state before destroying widgets
    current_state = get_current_widget_values(dynamic_frame)

    if dynamic_frame and dynamic_frame.winfo_exists():
        for widget in dynamic_frame.winfo_children():
            widget.destroy()
        # 2. Pass the saved state to the loading function
        load_all_dynamic_widgets(initial_state=current_state)

    # Refresh the constructor's listbox
    if listbox_to_refresh and listbox_to_refresh.winfo_exists():
        populate_tags_listbox(listbox_to_refresh)

    # Update Число values
    merge_data = get_common_merge_data()
    numbers = load_number_config()
    for num in numbers:
        val = evaluate_number_sequence(num['sequence'], merge_data)
        for widget in dynamic_frame.winfo_children():
            if widget._name == num['name']:
                widget.config(text=val)
    update_number_labels()



def export_all_tags_to_json(parent_window):
    """
    Silently collects all available tags, sorts them alphabetically,
    and saves them to a single JSON file. Only shows a message on error.
    """
    try:
        all_tags = set()

        # 1. Add the hardcoded 'today_tag'
        all_tags.add('today_tag')

        # 2. Load tags from fields_config.json
        fields_config = load_json(FIELDS_CONFIG_PATH, 'fields_config')
        for field in fields_config:
            all_tags.add(field['name'])

        # 3. Load tags from combobox_regular.json
        regular_combos = load_json(COMBOBOX_REGULAR_PATH, 'combobox_regular')
        for combo in regular_combos:
            all_tags.add(combo['name'])

        # 4. Load from combobox_mainkey.json (both main keys and all sub-keys)
        mainkey_combos = load_json(COMBOBOX_MAINKEY_PATH, 'combobox_mainkey')
        for combo in mainkey_combos:
            all_tags.add(combo['name'])  # Add the main key itself
            for mk_dict in combo.get('main_keys', []):
                if mk_dict:
                    subkeys_dict = list(mk_dict.values())[0]
                    for subkey_name in subkeys_dict.keys():
                        all_tags.add(subkey_name)

        # 5. Load from combination_config.json (only the combination's name)
        combination_config = load_json(COMBINATION_CONFIG_PATH, 'combination_config')
        for combo in combination_config:
            all_tags.add(combo['name'])

        number_config = load_number_config()
        for num in number_config:
            all_tags.add(num['name'])


        # Sort the final list alphabetically (case-insensitive)
        sorted_tags_list = sorted(list(all_tags), key=str.lower)

        # Save the sorted list to the new JSON file
        save_json(ALL_TAGS_OUTPUT_PATH, sorted_tags_list)

        # Print a silent confirmation to the console instead of a popup
        print(f"Tag list updated: {len(sorted_tags_list)} tags exported to {ALL_TAGS_OUTPUT_PATH}")

    except Exception as e:
        # Still show an error if something goes wrong
        messagebox.showerror("Ошибка экспорта тегов", f"Произошла ошибка при автоматическом экспорте тегов: {e}",
                             parent=parent_window)

def on_constructor_close(window_to_destroy):
    """Handles the closing event for the constructor window."""
    # The main 'window' is passed as the parent for any potential error message
    export_all_tags_to_json(window)
    window_to_destroy.destroy()

# --- Insert this helper somewhere near your other utility functions (after save_json) ---

def update_references(old_to_new: dict):
    """Replace tag names in rules_config and combination_config according to mapping {old: new}."""
    if not old_to_new:
        return
    try:
        # Update combination tags
        comb_cfg = load_json(COMBINATION_CONFIG_PATH, 'combination_config')
        comb_changed = False
        for combo in comb_cfg:
            tags = combo.get('tags', [])
            new_tags = [old_to_new.get(t, t) for t in tags]
            if new_tags != tags:
                combo['tags'] = new_tags
                comb_changed = True
        if comb_changed:
            save_json(COMBINATION_CONFIG_PATH, comb_cfg)

        # Update rules (conditions and behaviors)
        rules_cfg = load_json(RULES_CONFIG_PATH, 'rules_config')
        rules_changed = False
        for rule in rules_cfg:
            for cond in rule.get('conditions', []):
                t = cond.get('tag')
                if t in old_to_new:
                    cond['tag'] = old_to_new[t]
                    rules_changed = True
            for beh in rule.get('behaviors', []):
                t = beh.get('tag')
                if t in old_to_new:
                    beh['tag'] = old_to_new[t]
                    rules_changed = True
        if rules_changed:
            save_json(RULES_CONFIG_PATH, rules_cfg)

    except Exception as e:
        # don't crash the UI on small update problems — just log
        print(f"[update_references] error: {e}")


def get_common_merge_data():
    """Collects data from all dynamically created UI elements for the mail merge."""
    global dynamic_frame
    merge_data = {}

    # Add the only hardcoded tag value
    merge_data['today_tag'] = datetime.now().strftime('%d.%m.%Y')
    try:
        all_tags_from_file = load_json(ALL_TAGS_OUTPUT_PATH, 'all_tags.json')
        if isinstance(all_tags_from_file, list):
            string_tags = [str(tag) for tag in all_tags_from_file]
            merge_data['all_tags_merge'] = ', '.join(string_tags)
        else:
            merge_data['all_tags_merge'] = ''
    except Exception as e:
        print(f"Could not create 'all_tags_merge' due to an error: {e}")
        merge_data['all_tags_merge'] = ''

    # 1. Collect data from dynamically created widgets (Entry, Checkbox, etc.)
    if dynamic_frame:
        for widget in dynamic_frame.winfo_children():
            if not hasattr(widget, '_name'):
                continue
            widget_name = widget._name
            if isinstance(widget, (tk.Entry, tkentrycomplete.Combobox)):
                merge_data[widget_name] = widget.get()
            elif isinstance(widget, ttk.Checkbutton):
                var = checkbox_vars.get(widget_name)
                if var:
                    merge_data[widget_name] = "1" if var.get() else "0"

    # 2. Include raw subkeys from main_key combobox selections
    for data_dict in main_key_selections.values():
        merge_data.update(data_dict)

    # 3. MOVED UP: Calculate Число (Number) tags and add them to merge_data
    numbers = load_number_config()
    for num in numbers:
        val = evaluate_number_sequence(num['sequence'], merge_data)
        merge_data[num['name']] = val

    # 4. MOVED UP: Calculate Сочетание (Combination) tags and add them to merge_data
    combination_config = load_combination_config()
    for combo in combination_config:
        combined_value = ""
        for tag in combo['tags']:
            if tag == 'today_tag':
                combined_value += datetime.now().strftime("%d.%m.%Y")
            else:
                combined_value += merge_data.get(tag, tag)
        merge_data[combo['name']] = combined_value

    # 5. NOW, PERFORM DYNAMIC SUBKEY REPLACEMENT
    # This block now runs AFTER numbers and combinations are calculated and in merge_data.
    for main_key, sub_dict in main_key_selections.items():
        for subkey, value in sub_dict.items():
            if isinstance(value, str) and value.startswith('{') and value.endswith('}'):
                referenced_tag_name = value[1:-1]
                if referenced_tag_name in merge_data:
                    new_value = merge_data[referenced_tag_name]
                    sub_dict[subkey] = new_value
                    merge_data[subkey] = new_value

    # 6. Final formatting
    for key, val in merge_data.items():
        if isinstance(val, str):
            merge_data[key] = val.replace(" \\n ", "\n")

    return merge_data

def guess_type(value: str):
    """Try to guess if the string should be int, float, date, or text."""

    val = str(value).strip()

    # Try integer (only digits)
    if val.isdigit():
        try:
            return int(val)
        except ValueError:
            pass

    # Try float (handle decimal commas and points)
    try:
        if "." in val or "," in val:
            # Replace comma with dot for float conversion
            val_float = val.replace(",", ".")
            return float(val_float)
    except ValueError:
        pass

    # Try date with common formats
    date_formats = [
        "%Y-%m-%d",   # 2023-08-10
        "%d.%m.%Y",   # 10.08.2023
        "%d/%m/%Y",   # 10/08/2023
        "%m/%d/%Y",   # 08/10/2023
        "%d-%m-%Y",   # 10-08-2023
        "%Y/%m/%d",   # 2023/08/10
        # add more if needed
    ]

    for fmt in date_formats:
        try:
            dt = datetime.strptime(val, fmt)
            return dt.date()  # or dt if you want datetime
        except ValueError:
            continue

    # If all else fails, return original string
    return val

def submit_and_save():
    """Main function to generate documents after validation."""
    global dynamic_frame

    # Check if any fields have been created
    if not dynamic_frame or not dynamic_frame.winfo_children():
        messagebox.showinfo(
            "Информация",
            "Пожалуйста, создайте хотя бы одно поле в Конструкторе перед формированием документов."
        )
        return

    try:
        # Use global project-specific directories
        source_dir = TEMPLATES_DIR  # Points to documents/templates/<project_name>
        output_dir = PROCESSED_DIR  # Points to documents/processed/<project_name>

        # Ensure directories exist
        if not os.path.exists(source_dir):
            messagebox.showerror("Ошибка", f"Директория шаблонов '{source_dir}' не найдена. Проверьте конфигурацию проекта.")
            return
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        # Get merge data
        common_data = get_common_merge_data()

        # Apply rules
        rules = load_json(RULES_CONFIG_PATH, 'rules_config')
        if not isinstance(rules, list):
            rules = []

        for rule in rules:
            conditions = rule.get('conditions', [])
            behaviors = rule.get('behaviors', [])

            if not conditions:
                common_data = apply_behaviors(behaviors, common_data)
                continue

            all_conditions_met = all(evaluate_condition(cond, common_data) for cond in conditions)

            if all_conditions_met:
                non_cleaner_behaviors = [
                    b for b in behaviors if b.get('condition') != 'очистить при не выполнении'
                ]
                if non_cleaner_behaviors:
                    common_data = apply_behaviors(non_cleaner_behaviors, common_data)
            else:
                cleaner_behavior = next(
                    (b for b in behaviors if b.get('condition') == 'очистить при не выполнении'), None
                )
                if cleaner_behavior:
                    common_data = apply_behaviors([cleaner_behavior], common_data)

        # Timestamp for output filenames
        timestamp = datetime.now().strftime("%m%d%H%M%S")

        # Process DOCX files
        docx_files = [f for f in os.listdir(source_dir) if f.endswith('.docx')]
        for docx_file in docx_files:
            try:
                file_name_without_ext, _ = os.path.splitext(docx_file)
                new_file_name = f"{file_name_without_ext}_{timestamp}.docx"
                output_path = os.path.join(output_dir, new_file_name)

                if os.path.exists(output_path):
                    if not messagebox.askyesno(
                        "Перезапись файла",
                        f"Файл '{new_file_name}' уже существует. Хотите перезаписать?"
                    ):
                        continue

                document = MailMerge(os.path.join(source_dir, docx_file))
                merge_fields_in_doc = document.get_merge_fields()
                filtered_data = {key: common_data.get(key, '') for key in merge_fields_in_doc}

                document.merge(**filtered_data)
                document.write(output_path)
                document.close()
            except Exception as e:
                print(f"Error processing {docx_file}: {e}")

        # Process XLS/XLSX files
        xls_files = [f for f in os.listdir(source_dir) if f.endswith(('.xls', '.xlsx'))]
        for xls_file in xls_files:
            try:
                file_name_without_ext, _ = os.path.splitext(xls_file)
                new_file_name = f"{file_name_without_ext}_{timestamp}.xlsx"
                output_path = os.path.join(output_dir, new_file_name)

                if os.path.exists(output_path):
                    if not messagebox.askyesno(
                        "Перезапись файла",
                        f"Файл '{new_file_name}' уже существует. Хотите перезаписать?"
                    ):
                        continue

                wb = openpyxl.load_workbook(os.path.join(source_dir, xls_file))
                sheet = wb.active

                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and str(cell.value) in common_data:
                            new_val = guess_type(common_data[str(cell.value)])

                            # Assign type & number format
                            if isinstance(new_val, (int, float)):
                                cell.value = new_val
                                cell.number_format = numbers.FORMAT_GENERAL
                            elif isinstance(new_val, date):
                                cell.value = new_val
                                cell.number_format = 'DD.MM.YYYY'
                            else:
                                cell.value = str(new_val)
                                cell.number_format = numbers.FORMAT_GENERAL

                wb.save(output_path)

            except Exception as e:
                print(f"Error processing {xls_file}: {e}")

        messagebox.showinfo(title="Успех!", message="Документы были успешно сформированы.")

    except Exception as e:
        messagebox.showerror(
            title="Ошибка!",
            message=f"Произошла ошибка при формировании документов: {e}"
        )

def clear_all_inputs():
    """Reset all entryboxes, checkboxes, and comboboxes inside dynamic_frame."""
    global dynamic_frame, checkbox_vars, main_key_selections, ALL_TAG_VALUES

    if not dynamic_frame:
        return

    for widget in dynamic_frame.winfo_children():
        if isinstance(widget, tk.Entry):
            widget.delete(0, tk.END)
        elif isinstance(widget, ttk.Combobox):
            widget.set("")
        elif isinstance(widget, ttk.Checkbutton):
            var = checkbox_vars.get(widget._name)
            if var:
                var.set(0)

    # Also clear global state
    ALL_TAG_VALUES.clear()
    main_key_selections.clear()


def on_closing():
    """Handle closing of the main window with option to keep or erase data."""
    answer = messagebox.askyesnocancel(
        "Выход",
        "Вы хотите сохранить введённые данные?\n"
        "Да = оставить данные на месте\n"
        "Нет = очистить все данные\n"
        "Отмена = вернуться"
    )
    if answer is None:  # Cancel
        return
    elif answer:  # Yes → Save and keep data
        save_input_state()
        window.destroy()
    else:  # No → Erase all inputs and clear JSON
        clear_all_inputs()
        save_json(INPUT_STATE_PATH, {})  # overwrite with empty JSON
        window.destroy()

def prompt_before_switch():
    """Ask user whether to keep or clear data when switching projects."""
    answer = messagebox.askyesnocancel(
        "Переключение проекта",
        "Вы хотите сохранить введённые данные перед переключением?\n"
        "Да = сохранить данные\n"
        "Нет = очистить все данные\n"
        "Отмена = вернуться"
    )
    if answer is None:  # Cancel
        return False
    elif answer:  # Yes → Save
        save_input_state()
        return True
    else:  # No → Clear
        try:
            clear_all_inputs()
        except Exception:
            pass
        save_json(INPUT_STATE_PATH, {})
        return True

def import_fields():
    global dynamic_frame

    # Prompt user for confirmation
    if not messagebox.askyesno("Подтверждение", "Вы уверены, что хотите импортировать поля из файла 'field_import'?"):
        return

    # Define file paths relative to the import folder
    xlsx_path = os.path.join(IMPORT_FLD, "field_import.xlsx")
    xls_path = os.path.join(IMPORT_FLD, "field_import.xls")

    # Check for file existence, prioritizing .xlsx
    if os.path.exists(xlsx_path):
        file_path = xlsx_path
    elif os.path.exists(xls_path):
        file_path = xls_path
    else:
        messagebox.showerror("Ошибка", "Ошибка: Файл 'field_import' не найден.")
        return

    try:
        # Load Excel file
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        # Validate at least two columns (name + value)
        if sheet.max_column < 2:
            messagebox.showerror("Ошибка",
                                 "Ошибка: Файл 'field_import' поврежден или не содержит как минимум двух столбцов (имя, значение).")
            return

        # Initialize counters and tracking
        fields_created = 0
        values_imported = 0
        processed_names = set()  # Track field names (case-insensitive)
        value_map = {}  # Store values for reapplying after refresh

        # Load all config files
        fields_config = load_json(FIELDS_CONFIG_PATH, [])
        combobox_regular = load_json(COMBOBOX_REGULAR_PATH, [])
        combobox_mainkey = load_json(COMBOBOX_MAINKEY_PATH, [])
        combination_config = load_json(COMBINATION_CONFIG_PATH, [])

        # Process Excel rows
        for row in sheet.iter_rows(values_only=True):
            field_name = row[0]
            field_value = row[1]
            field_type = row[2] if len(row) >= 3 else None  # optional third column

            # Skip rows with empty field name
            if not field_name or not str(field_name).strip():
                continue

            # Normalize field type
            if not field_type or str(field_type).strip() == "":
                field_type = "текст"  # default
            else:
                field_type = str(field_type).strip().lower()
                if field_type not in ["текст", "числа", "дата"]:
                    messagebox.showerror(
                        "Ошибка",
                        f"Ошибка: Недопустимый тип '{field_type}' для поля '{field_name}'. "
                        f"Разрешенные типы: текст, числа, дата."
                    )
                    return  # cancel whole import

            # Check for duplicate field name (case-insensitive)
            field_name_lower = str(field_name).lower()
            if field_name_lower in processed_names:
                continue
            processed_names.add(field_name_lower)

            # Store value for later use (even if empty)
            value_map[field_name_lower] = str(field_value) if field_value is not None else ""

            # Check if field exists in any config (case-insensitive)
            exists = False
            tag_type = None
            for config in [fields_config, combobox_regular, combobox_mainkey, combination_config]:
                for item in config:
                    if item.get("name", "").lower() == field_name_lower:
                        exists = True
                        tag_type = item.get("tag_type")
                        break
                if exists:
                    break

            if exists:
                # If field exists and is type "поле", update its value
                if tag_type == "поле":
                    found = False
                    for widget in dynamic_frame.winfo_children():
                        if hasattr(widget, "_name") and widget._name.lower() == field_name_lower and isinstance(widget, tk.Entry):
                            widget.delete(0, tk.END)
                            widget.insert(0, value_map[field_name_lower])
                            values_imported += 1
                            found = True
                            break
                    if not found:
                        print(f"Warning: No Entry widget found for existing field '{field_name}'")
                continue

            # Create new field if unique
            fields_config.append({
                "name": str(field_name),
                "type": field_type,
                "tag_type": "поле"
            })
            add_dynamic_widget(str(field_name), field_type, "поле", value_map[field_name_lower])
            fields_created += 1
            values_imported += 1

        # Save updated fields config
        save_json(FIELDS_CONFIG_PATH, fields_config)

        # Refresh UI and reapply values
        refresh_main_and_constructor()

        # Reapply values to all Entry widgets
        for widget in dynamic_frame.winfo_children():
            if hasattr(widget, "_name") and isinstance(widget, tk.Entry):
                field_name_lower = widget._name.lower()
                if field_name_lower in value_map:
                    widget.delete(0, tk.END)
                    widget.insert(0, value_map[field_name_lower])

        # Show feedback
        messagebox.showinfo(
            "Импорт завершен",
            f"{fields_created} полей было импортировано, {values_imported} значений было импортировано."
        )

    except Exception as e:
        messagebox.showerror(
            "Ошибка",
            "Ошибка: Файл 'field_import' поврежден или не содержит необходимых столбцов (имя, значение, тип)."
        )
        print(f"Import error: {str(e)}")



def get_next_grid_position():
    """Calculates the next grid position (row, column) based on a 30-widget-per-column rule."""
    global dynamic_frame
    if not dynamic_frame:
        return 0, 0

    # Each widget consists of a Label and an Entry/Combobox, so we count pairs.
    widget_count = len(dynamic_frame.winfo_children()) // 2

    row = widget_count % 30
    col_group = widget_count // 30

    # Each column group for widgets takes 2 grid columns (one for label, one for widget)
    base_col = col_group * 2

    return row, base_col


def add_dynamic_widget(name, data_type, tag_type, values=None, main_key_data=None, initial_value=None):
    """Adds a new widget to the dynamic_frame, optionally with an initial value.
       Wires ALL_TAG_VALUES so Число labels update on the fly.
    """
    global dynamic_frame, checkbox_vars, main_key_selections, ALL_TAG_VALUES, NUMBER_LABELS

    row, base_col = get_next_grid_position()

    # For 'число' we create its own paired labels below; skip the generic label here
    if tag_type != "число":
        label = tk.Label(dynamic_frame, text=f"{name}:")
        label._name = f"{name}l"
        label.grid(row=row, column=base_col, padx=5, pady=2, sticky="e")

    if tag_type == "поле":
        var = tk.StringVar()
        if initial_value is not None:
            var.set(str(initial_value))
        entry = tk.Entry(dynamic_frame, textvariable=var, width=25)
        entry._name = name
        entry.grid(row=row, column=base_col + 1, padx=5, pady=2, sticky="w")

        ALL_TAG_VALUES[name] = var.get()
        var.trace_add("write", lambda *args, v=var, n=name: (ALL_TAG_VALUES.__setitem__(n, v.get()), update_number_labels()))

    elif tag_type == "чекбокс":
        var = tk.IntVar()
        if initial_value is not None:
            var.set(initial_value)
        checkbox_vars[name] = var
        checkbox = ttk.Checkbutton(dynamic_frame, variable=var)
        checkbox._name = name
        checkbox.grid(row=row, column=base_col + 1, padx=5, pady=2, sticky="w")

        ALL_TAG_VALUES[name] = var.get()
        var.trace_add("write", lambda *args, v=var, n=name: (ALL_TAG_VALUES.__setitem__(n, v.get()), update_number_labels()))

    elif tag_type == "комбобокс":  # Regular combobox
        var = tk.StringVar()
        combobox = tkentrycomplete.Combobox(dynamic_frame, values=values, textvariable=var, width=22)
        combobox._name = name
        if initial_value is not None:
            var.set(initial_value)
        combobox.set_completion_list({v: {} for v in values})
        combobox.grid(row=row, column=base_col + 1, padx=5, pady=2, sticky="w")

        ALL_TAG_VALUES[name] = var.get()
        combobox.bind("<<ComboboxSelected>>", lambda e, v=var, n=name: (ALL_TAG_VALUES.__setitem__(n, v.get()), update_number_labels()))
        var.trace_add("write", lambda *args, v=var, n=name: (ALL_TAG_VALUES.__setitem__(n, v.get()), update_number_labels()))

    elif tag_type == "список":  # Main-key combobox
        var = tk.StringVar()
        if initial_value is not None:
            var.set(initial_value)
        combobox = tkentrycomplete.Combobox(dynamic_frame, values=values, textvariable=var, width=22)
        combobox._name = name
        combobox.set_completion_list(main_key_data)

        def on_select(event=None, widget_name=name, data=main_key_data, v=var):
            selected_key = v.get()
            if selected_key in data:
                main_key_selections[widget_name] = data[selected_key]
            elif widget_name in main_key_selections:
                del main_key_selections[widget_name]
            ALL_TAG_VALUES[widget_name] = v.get()
            update_number_labels()

        combobox.bind('<<ComboboxSelected>>', on_select)
        combobox.bind('<FocusOut>', on_select)
        combobox.bind('<Return>', on_select)
        combobox.grid(row=row, column=base_col + 1, padx=5, pady=2, sticky="w")

        ALL_TAG_VALUES[name] = var.get()
        var.trace_add("write", lambda *args, v=var, n=name: (ALL_TAG_VALUES.__setitem__(n, v.get()), update_number_labels()))

    elif tag_type == "число":
        # Name label + value label
        label_name = tk.Label(dynamic_frame, text=f"{name}:")
        label_name._name = f"{name}l"
        label_name.grid(row=row, column=base_col, padx=5, pady=2, sticky="e")

        value_label = tk.Label(dynamic_frame, text="0,00")  # default display
        value_label._name = name
        value_label.grid(row=row, column=base_col + 1, padx=5, pady=2, sticky="w")

        NUMBER_LABELS[name] = value_label  # keep reference for live updates



def load_all_dynamic_widgets(initial_state=None):
    """Loads all configured UI elements, optionally applying an initial state."""

    NUMBER_LABELS.clear()

    # If no state passed in, load from saved JSON
    if initial_state is None:
        initial_state = load_input_state()

    fields_list = []
    comboboxes_list = []
    checkboxes_list = []
    numbers_list = []   # <-- NEW group for Число

    # --- Load Fields & Checkboxes ---
    fields_config = load_json(FIELDS_CONFIG_PATH, 'fields_config')
    for field in fields_config:
        tag_type = field.get('tag_type', 'поле')
        if tag_type == 'чекбокс':
            checkboxes_list.append((field['name'], field['type'], tag_type, None, None))
        else:  # Regular text field
            fields_list.append((field['name'], field['type'], tag_type, None, None))

    # --- Load Regular Comboboxes ---
    regular_combos = load_json(COMBOBOX_REGULAR_PATH, 'combobox_regular')
    for combo in regular_combos:
        comboboxes_list.append((combo['name'], combo['type'], 'комбобокс', combo['values'], None))

    # --- Load Main-Key Comboboxes ---
    mainkey_combos = load_json(COMBOBOX_MAINKEY_PATH, 'combobox_mainkey')
    for combo in mainkey_combos:
        values = [list(mk.keys())[0] for mk in combo['main_keys']]
        data_dict = {list(mk.keys())[0]: list(mk.values())[0] for mk in combo['main_keys']}
        comboboxes_list.append((combo['name'], combo['type'], 'список', values, data_dict))

    # --- Load Число tags ---
    numbers = load_number_config()
    for num in numbers:
        numbers_list.append((num['name'], None, 'число', None, None))

    # --- Sort each group alphabetically by name ---
    fields_list.sort(key=lambda x: x[0].lower())
    comboboxes_list.sort(key=lambda x: x[0].lower())
    checkboxes_list.sort(key=lambda x: x[0].lower())
    numbers_list.sort(key=lambda x: x[0].lower())   # <-- sort числа too

    # --- Combine groups in desired order ---
    ordered_widgets = fields_list + comboboxes_list + checkboxes_list + numbers_list

    # --- Place widgets in order, passing the saved value ---
    for idx, (name, data_type, tag_type, values, main_key_data) in enumerate(ordered_widgets):
        saved_value = initial_state.get(name)
        add_dynamic_widget(name, data_type, tag_type, values, main_key_data, initial_value=saved_value)
    update_number_labels()


def get_all_tags_for_constructor():
    """Gathers a flat list of all tags for use in constructor UI elements."""
    all_tags = []

    # Add fields, checkboxes, and regular comboboxes
    all_tags.extend([f['name'] for f in load_json(FIELDS_CONFIG_PATH, 'fields_config')])
    all_tags.extend([c['name'] for c in load_json(COMBOBOX_REGULAR_PATH, 'combobox_regular')])

    # Add main keys and their subkeys
    mainkey_combos = load_json(COMBOBOX_MAINKEY_PATH, 'combobox_mainkey')
    for combo in mainkey_combos:
        # Add the main key itself
        all_tags.append(combo['name'])
        # Add each subkey
        for mk_dict in combo.get('main_keys', []):
            # Assumes each main_keys item is a dictionary with one key (main_key)
            subkeys_dict = list(mk_dict.values())[0]
            for subkey_name in subkeys_dict.keys():
                all_tags.append(subkey_name)

    # Add combination tags
    combination_config = load_json(COMBINATION_CONFIG_PATH, 'combination_config')
    all_tags.extend([c['name'] for c in combination_config])

    number_config = load_number_config()
    all_tags.extend([n['name'] for n in number_config])

    return sorted(list(set(all_tags)))  # Return sorted unique tags

# --- Constructor Window and its Helpers ---

def sort_column(treeview, col, reverse):
    """Sorts a Treeview column."""
    items = [(treeview.set(item, col), item) for item in treeview.get_children('')]
    items.sort(key=lambda x: str(x[0]).lower(), reverse=reverse)
    for index, (_, item) in enumerate(items):
        treeview.move(item, '', index)
    treeview.heading(col, command=lambda: sort_column(treeview, col, not reverse))


def update_rules_listbox(rules, listbox):
    # Calculate maximum number of conditions and behaviors, with minimum 1 each
    max_conditions = max((len(rule.get('conditions', [])) for rule in rules), default=1) if rules else 1
    max_behaviors = max((len(rule.get('behaviors', [])) for rule in rules), default=1) if rules else 1

    # Define columns: always include "Имя", "Условие 1", "Поведение 1"
    required_columns = ["Имя"] + [f"Условие {i + 1}" for i in range(max_conditions)] + [f"Поведение {i + 1}" for i in
                                                                                           range(max_behaviors)]

    # Always set columns to ensure default columns are present
    listbox["columns"] = required_columns
    for col in required_columns:
        listbox.heading(col, text=col, anchor="center")
        listbox.column(col, anchor="center", stretch=True)

    # Clear existing items
    for item in listbox.get_children():
        listbox.delete(item)

    # Populate listbox
    for rule in rules:
        name = rule.get('name', '')
        conditions = [f"{c['tag']} {c['condition']} {c['rule']}" for c in rule.get('conditions', [])]
        behaviors = [f"{b['tag']} {b['condition']} {b['rule']}" for b in rule.get('behaviors', [])]
        # Align values with columns: name, conditions, padding, behaviors, padding
        values = [name] + conditions + [""] * (max_conditions - len(conditions)) + behaviors + [""] * (
            max_behaviors - len(behaviors))
        listbox.insert("", "end", values=values)


def open_constructor_window():
    """Opens the main constructor window for managing tags and rules."""
    constructor_window = tk.Toplevel(window)
    constructor_window.title("Конструктор")
    constructor_window.geometry("1200x600")
    constructor_window.focus_set()

    # This intercepts the "X" button press and calls our custom function
    constructor_window.protocol("WM_DELETE_WINDOW", lambda: on_constructor_close(constructor_window))

    notebook = ttk.Notebook(constructor_window)
    notebook.pack(pady=10, padx=10, fill="both", expand=True)

    # --- Tags Tab ---
    tags_tab = ttk.Frame(notebook)
    notebook.add(tags_tab, text='Теги')

    tags_list_frame = tk.Frame(tags_tab)
    tags_list_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
    tags_buttons_frame = tk.Frame(tags_tab)
    tags_buttons_frame.grid(row=0, column=1, padx=5, pady=5, sticky="ns")
    tags_tab.grid_rowconfigure(0, weight=1)
    tags_tab.grid_columnconfigure(0, weight=1)

    tags_listbox = ttk.Treeview(tags_list_frame, columns=("Имя", "Тип ввода", "Тип тега"), show="headings", height=20)
    tags_listbox.heading("Имя", text="Имя", command=lambda: sort_column(tags_listbox, "Имя", False))
    tags_listbox.heading("Тип ввода", text="Тип ввода", command=lambda: sort_column(tags_listbox, "Тип ввода", False))
    tags_listbox.heading("Тип тега", text="Тип тега", command=lambda: sort_column(tags_listbox, "Тип тега", False))
    tags_listbox.column("Имя", width=250, anchor="center")
    tags_listbox.column("Тип ввода", width=120, anchor="center")
    tags_listbox.column("Тип тега", width=120, anchor="center")

    scrollbar_tags = ttk.Scrollbar(tags_list_frame, orient="vertical", command=tags_listbox.yview)
    tags_listbox.configure(yscrollcommand=scrollbar_tags.set)
    tags_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar_tags.pack(side=tk.RIGHT, fill=tk.Y)

    tk.Button(tags_buttons_frame, text="Новый", width=14, command=lambda: open_new_tag_window(tags_listbox, constructor_window)).pack(
        side=TOP, pady=2)
    tk.Button(tags_buttons_frame, text="Редактировать", width=14,
              command=lambda: open_edit_tag_window(tags_listbox, constructor_window)).pack(side=TOP, pady=2)
    tk.Button(tags_buttons_frame, text="Удалить", width=14, command=lambda: delete_tag(tags_listbox, constructor_window)).pack(side=TOP,
                                                                                                           pady=2)
    populate_tags_listbox(tags_listbox)

    # --- Rules Tab ---
    rules_tab = ttk.Frame(notebook)
    notebook.add(rules_tab, text='Правила')

    rules_list_frame = tk.Frame(rules_tab)
    rules_list_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
    rules_buttons_frame = tk.Frame(rules_tab)
    rules_buttons_frame.grid(row=0, column=1, padx=5, pady=5, sticky="ns")
    rules_tab.grid_rowconfigure(0, weight=1)
    rules_tab.grid_columnconfigure(0, weight=1)

    rules_listbox = ttk.Treeview(rules_list_frame, show="headings", height=20)

    v_scrollbar_rules = ttk.Scrollbar(rules_list_frame, orient="vertical", command=rules_listbox.yview)
    h_scrollbar_rules = ttk.Scrollbar(rules_list_frame, orient="horizontal", command=rules_listbox.xview)
    rules_listbox.configure(yscrollcommand=v_scrollbar_rules.set, xscrollcommand=h_scrollbar_rules.set)

    rules_listbox.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
    v_scrollbar_rules.pack(side=tk.RIGHT, fill=tk.Y, before=rules_listbox)
    h_scrollbar_rules.pack(side=tk.BOTTOM, fill=tk.X)

    tk.Button(rules_buttons_frame, text="Создать", width=14,
              command=lambda: open_create_rule_window(rules_listbox, constructor_window)).pack(side=TOP, pady=2)
    tk.Button(rules_buttons_frame, text="Изменить", width=14,
              command=lambda: open_edit_rule_window(rules_listbox, constructor_window)).pack(side=TOP, pady=2)
    tk.Button(rules_buttons_frame, text="Удалить", width=14, command=lambda: delete_rule(rules_listbox, constructor_window)).pack(side=TOP,
                                                                                                              pady=2)

    rules = load_json(RULES_CONFIG_PATH, 'rules_config')
    update_rules_listbox(rules, rules_listbox)


# All other constructor helper functions (open_new_tag_window, open_field_window, etc.) need to be included here.
# For brevity, I will add the main ones back. The complex ones like rules will need to be added from original.
# The following implementations are from the original file, adapted for the new structure.

def open_new_tag_window(listbox, parent_window):
    """Window to choose what kind of new tag to create."""
    new_window = tk.Toplevel(parent_window)
    new_window.title("Новый тег")
    new_window.wm_minsize(width=250, height=0)
    new_window.resizable(False, False)
    new_window.focus_set()
    new_window.grab_set()

    btn_frame = tk.Frame(new_window)
    btn_frame.pack(pady=10, expand=True)

    # The listbox needs to be passed to refresh it upon creation
    tk.Button(btn_frame, text="ПОЛЕ", width=15, height=2, command=lambda: [new_window.destroy(), open_field_window(listbox, None, parent_window)]).pack(pady=5)
    tk.Button(btn_frame, text="СПИСОК", width=15, height=2,
              command=lambda: [new_window.destroy(), open_list_window(listbox, None, parent_window)]).pack(pady=5)
    tk.Button(btn_frame, text="ЧЕКБОКС", width=15, height=2,
              command=lambda: [new_window.destroy(), open_checkbox_window(listbox, None, parent_window)]).pack(pady=5)
    tk.Button(btn_frame, text="СОЧЕТАНИЕ", width=15, height=2, command=lambda: [new_window.destroy(), open_combination_window(listbox, None, parent_window)]).pack(pady=5)
    tk.Button(btn_frame, text="ЧИСЛО", width=15, height=2, command=lambda: [new_window.destroy(), open_number_window(listbox, None, parent_window)]).pack(pady=5)


def open_field_window(listbox, item_to_edit, parent_window):
    """Window to create or edit a simple 'Field' (Entry widget)."""
    is_edit = item_to_edit is not None
    title = "Редактировать поле" if is_edit else "Создание поля"

    field_window = tk.Toplevel(parent_window)
    field_window.title(title)
    field_window.geometry("300x180")
    field_window.resizable(False, False)
    field_window.focus_set()
    field_window.grab_set()

    old_name = ""
    if is_edit:
        old_name, old_type, _ = listbox.item(item_to_edit)['values']
        name_var = tk.StringVar(value=old_name)
        type_var = tk.StringVar(value=old_type)
    else:
        name_var = tk.StringVar()
        type_var = tk.StringVar(value="текст")

    tk.Label(field_window, text="Имя поля:").pack(pady=(10, 0))
    name_entry = tk.Entry(field_window, textvariable=name_var, width=30)
    name_entry.pack(pady=5, padx=10)
    name_entry.focus()

    tk.Label(field_window, text="Тип данных:").pack(pady=(10, 0))
    type_frame = tk.Frame(field_window)
    type_frame.pack(pady=5)
    types = [("текст", "текст"), ("числа", "числа"), ("дата", "дата")]
    for text, value in types:
        tk.Radiobutton(type_frame, text=text, value=value, variable=type_var).pack(side=LEFT, padx=5)

    def save_field():
        name = name_var.get().strip()
        data_type = type_var.get()
        if not name:
            messagebox.showwarning("Ошибка", "Введите имя поля.", parent=field_window)
            return

        all_configs = (load_json(path, '') for path in
                       [FIELDS_CONFIG_PATH, COMBOBOX_REGULAR_PATH, COMBOBOX_MAINKEY_PATH, COMBINATION_CONFIG_PATH])
        all_names = {item['name'] for config in all_configs for item in config}

        if name != old_name and name in all_names:
            messagebox.showwarning("Ошибка", f"Имя '{name}' уже используется.", parent=field_window)
            return

        fields_config = load_json(FIELDS_CONFIG_PATH, 'fields_config')
        if is_edit:
            for field in fields_config:
                if field['name'] == old_name:
                    field['name'] = name
                    field['type'] = data_type
                    break
        else:
            fields_config.append({'name': name, 'type': data_type, 'tag_type': 'поле'})

        save_json(FIELDS_CONFIG_PATH, fields_config)
        if is_edit and old_name != name:
            update_references({old_name: name})
        refresh_all_windows(listbox)
        field_window.destroy()

    btn_frame = tk.Frame(field_window)
    btn_frame.pack(pady=10)
    tk.Button(btn_frame, text="ОК", width=10, command=save_field).pack(side=LEFT, padx=5)
    tk.Button(btn_frame, text="ОТМЕНА", width=10, command=field_window.destroy).pack(side=LEFT, padx=5)


def open_checkbox_window(listbox, item_to_edit, parent_window):
    """Window to create or edit a 'Checkbox'."""
    is_edit = item_to_edit is not None
    title = "Редактировать чекбокс" if is_edit else "Создание чекбокса"

    cb_window = tk.Toplevel(parent_window)
    cb_window.title(title)
    cb_window.geometry("300x120")
    cb_window.resizable(False, False)
    cb_window.focus_set()
    cb_window.grab_set()

    old_name = ""
    if is_edit:
        old_name, _, _ = listbox.item(item_to_edit)['values']
        name_var = tk.StringVar(value=old_name)
    else:
        name_var = tk.StringVar()

    tk.Label(cb_window, text="Имя чекбокса:").pack(pady=(10, 0))
    name_entry = tk.Entry(cb_window, textvariable=name_var, width=30)
    name_entry.pack(pady=5, padx=10)
    name_entry.focus()

    def save_checkbox():
        name = name_var.get().strip()
        if not name:
            messagebox.showwarning("Ошибка", "Введите имя чекбокса.", parent=cb_window)
            return

        all_configs = (load_json(path, '') for path in
                       [FIELDS_CONFIG_PATH, COMBOBOX_REGULAR_PATH, COMBOBOX_MAINKEY_PATH, COMBINATION_CONFIG_PATH])
        all_names = {item['name'] for config in all_configs for item in config}

        if name != old_name and name in all_names:
            messagebox.showwarning("Ошибка", f"Имя '{name}' уже используется.", parent=cb_window)
            return

        fields_config = load_json(FIELDS_CONFIG_PATH, 'fields_config')
        if is_edit:
            for field in fields_config:
                if field['name'] == old_name:
                    field['name'] = name
                    break
        else:
            fields_config.append({'name': name, 'type': 'чекбокс', 'tag_type': 'чекбокс'})

        save_json(FIELDS_CONFIG_PATH, fields_config)
        if is_edit and old_name != name:
            update_references({old_name: name})
        refresh_all_windows(listbox)
        cb_window.destroy()

    btn_frame = tk.Frame(cb_window)
    btn_frame.pack(pady=10)
    tk.Button(btn_frame, text="ОК", width=10, command=save_checkbox).pack(side=LEFT, padx=5)
    tk.Button(btn_frame, text="ОТМЕНА", width=10, command=cb_window.destroy).pack(side=LEFT, padx=5)


# --- REPLACE THE OLD open_list_window FUNCTION WITH THIS ---

def open_list_window(listbox, item_to_edit, parent_window,constructor_listbox_ref=None):
    """
    Full implementation for CREATING and EDITING lists.
    Now for 'список' type in edit mode, uses a main key selector instead of showing all main keys at once.
    """
    is_edit = item_to_edit is not None
    title = "Редактировать список" if is_edit else "Создание списка"

    list_window = tk.Toplevel(parent_window)
    list_window.title(title)
    list_window.resizable(False, False)
    list_window.focus_set()
    list_window.grab_set()

    # --- Initial Variable Setup ---
    name_var = tk.StringVar()
    main_key_var = tk.IntVar()
    sets_list = []
    old_name = ""
    original_tag_type = ""

    # --- Data Loading for Edit Mode ---
    if is_edit:
        old_name, _, tag_type = listbox.item(item_to_edit)['values']
        original_tag_type = tag_type
        name_var.set(old_name)

        if tag_type == 'комбобокс':
            main_key_var.set(0)
            config = load_json(COMBOBOX_REGULAR_PATH, '')
            list_data = next((item for item in config if item['name'] == old_name), None)
            if list_data:
                sets_list.append({'initial_values': list_data.get('values', [])})

        elif tag_type == 'список':
            main_key_var.set(1)
            config = load_json(COMBOBOX_MAINKEY_PATH, '')
            list_data = next((item for item in config if item['name'] == old_name), None)
            if list_data:
                for main_key_dict in list_data.get('main_keys', []):
                    for main_key, sub_dict in main_key_dict.items():
                        new_set = {
                            "main_key": tk.StringVar(value=main_key),
                            "key_values": [{"key": tk.StringVar(value=k), "value": tk.StringVar(value=v)}
                                           for k, v in sub_dict.items()]
                        }
                        sets_list.append(new_set)
    else:
        sets_list.append({"main_key": tk.StringVar(), "key_values": [{"key": tk.StringVar(), "value": tk.StringVar()}]})

    # --- Helper Functions ---
    def add_key_value_row():
        for s in sets_list:
            s["key_values"].append({"key": tk.StringVar(), "value": tk.StringVar()})
        refresh_table()

    def add_key_value_to_specific_set(set_index):
        if 0 <= set_index < len(sets_list):
            sets_list[set_index]["key_values"].append({"key": tk.StringVar(), "value": tk.StringVar()})
            refresh_table()

    def add_set():
        # Get subkey names from the first set (ignores values, just keeps keys)
        if sets_list:
            key_structure = [kv['key'].get() for kv in sets_list[0]['key_values']]
        else:
            key_structure = []

        # Create new key-value dicts: same keys, empty values
        new_key_values = [{"key": tk.StringVar(value=k), "value": tk.StringVar()} for k in key_structure]

        sets_list.append({
            "main_key": tk.StringVar(),
            "key_values": new_key_values
        })

        refresh_table()

    def refresh_table():
        for widget in table_frame.winfo_children():
            widget.destroy()

        if not main_key_var.get():
            # --- Simple list mode ---
            simple_frame = tk.Frame(table_frame)
            simple_frame.pack(fill="both", expand=True)
            tk.Label(simple_frame, text="Значения (каждое с новой строки):").pack(anchor="w")
            text_area = Text(simple_frame, width=60, height=15)
            text_area.pack(fill="both", expand=True, padx=5, pady=5)
            if is_edit and sets_list and 'initial_values' in sets_list[0]:
                text_area.insert('1.0', '\n'.join(sets_list[0]['initial_values']))
            if not is_edit:
                tk.Button(simple_frame, text="Импорт из Excel", command=import_from_excel).pack(pady=5)
            sets_list[0]['widget_ref'] = text_area



        else:

            if is_edit:

                # --- Main key edit mode with selector ---

                selector_frame = tk.Frame(table_frame)

                selector_frame.pack(fill="x", padx=10, pady=5)

                tk.Label(selector_frame, text="Выберите главный ключ:", anchor="w").pack(side="left")

                main_keys_list = [s["main_key"].get() for s in sets_list if s["main_key"].get().strip()]

                selected_main_key = tk.StringVar()

                main_key_combo = ttk.Combobox(selector_frame, values=main_keys_list,

                                              textvariable=selected_main_key, state="readonly", width=25)

                main_key_combo.pack(side="left", padx=5)

                subkeys_frame = tk.Frame(table_frame)

                subkeys_frame.pack(fill="x", padx=10, pady=5)

                def show_subkeys_for_selected(*args):

                    for w in subkeys_frame.winfo_children():
                        w.destroy()

                    key = selected_main_key.get()

                    set_index = next((i for i, s in enumerate(sets_list) if s["main_key"].get() == key), None)

                    if set_index is None:
                        return

                    s = sets_list[set_index]

                    tk.Label(subkeys_frame, text=f"Главный ключ:", anchor="w").grid(row=0, column=0, sticky="w", pady=2)

                    tk.Entry(subkeys_frame, textvariable=s["main_key"], width=20).grid(row=0, column=1, sticky="w",
                                                                                       padx=5)

                    for i, kv in enumerate(s["key_values"], start=1):
                        tk.Label(subkeys_frame, text=f"Ключ {i}:", anchor="w").grid(row=i, column=0, sticky="w", pady=2)

                        tk.Entry(subkeys_frame, textvariable=kv["key"], width=20).grid(row=i, column=1, sticky="w",
                                                                                       padx=5)

                        tk.Label(subkeys_frame, text="Значение:", anchor="w").grid(row=i, column=2, sticky="w", padx=5)

                        tk.Entry(subkeys_frame, textvariable=kv["value"], width=20).grid(row=i, column=3, sticky="w",
                                                                                         padx=5)

                    tk.Button(subkeys_frame, text="Добавить строку",

                              command=lambda: add_key_value_to_specific_set(set_index)).grid(

                        row=len(s["key_values"]) + 1, column=0, pady=5)

                main_key_combo.bind("<<ComboboxSelected>>", show_subkeys_for_selected)

                if main_keys_list:
                    selected_main_key.set(main_keys_list[0])

                    show_subkeys_for_selected()

                control_frame = tk.Frame(table_frame)

                control_frame.pack(pady=10)

                def add_set_and_update_dropdown():
                    # Get subkey names from the first set (ignores values, keeps keys)
                    if sets_list:
                        key_structure = [kv['key'].get() for kv in sets_list[0]['key_values']]
                    else:
                        key_structure = []

                    # Create new set with empty main key name
                    new_key_values = [{"key": tk.StringVar(value=k), "value": tk.StringVar()} for k in key_structure]
                    empty_main_key_var = tk.StringVar(value="")
                    sets_list.append({
                        "main_key": empty_main_key_var,
                        "key_values": new_key_values
                    })

                    # Update dropdown list with empty name (will appear as blank)
                    main_keys_list.append("")
                    main_key_combo['values'] = main_keys_list

                    # Select the new (blank) main key in dropdown
                    selected_main_key.set("")
                    show_subkeys_for_selected()
                tk.Button(control_frame, text="Добавить главный ключ", command=add_set_and_update_dropdown).pack(side=LEFT, padx=5)



            else:

                # --- Main key creation mode ---

                for set_index, s in enumerate(sets_list):

                    set_frame = ttk.LabelFrame(table_frame, text=f"Набор {set_index + 1}")

                    set_frame.pack(fill="x", expand=True, padx=5, pady=5)

                    keys_frame = tk.Frame(set_frame)

                    keys_frame.pack(fill="x", padx=10, pady=2)

                    tk.Label(keys_frame, text="Главный ключ:", anchor="w").grid(row=0, column=0, sticky="w", pady=2)

                    tk.Entry(keys_frame, textvariable=s["main_key"], width=25).grid(row=0, column=1, sticky="w", padx=5)

                    for i, kv in enumerate(s["key_values"]):
                        row_num = i + 1

                        tk.Label(keys_frame, text=f"Ключ {row_num}:", anchor="w").grid(row=row_num, column=0,
                                                                                       sticky="w", pady=2)

                        tk.Entry(keys_frame, textvariable=kv["key"], width=20).grid(row=row_num, column=1, sticky="w",
                                                                                    padx=5)

                        tk.Label(keys_frame, text="Значение:", anchor="w").grid(row=row_num, column=2, sticky="w",
                                                                                padx=5)

                        tk.Entry(keys_frame, textvariable=kv["value"], width=20).grid(row=row_num, column=3, sticky="w",
                                                                                      padx=5)

                    tk.Button(set_frame, text="Добавить строку в этот набор",

                              command=lambda si=set_index: add_key_value_to_specific_set(si)).pack(pady=5)

                control_frame = tk.Frame(table_frame)

                control_frame.pack(pady=10)

                tk.Button(control_frame, text="Строка", command=add_key_value_row).pack(side=LEFT, padx=5)

                tk.Button(control_frame, text="Добавить", command=add_set).pack(side=LEFT, padx=5)

                if not is_edit:
                    tk.Button(control_frame, text="Импорт", command=import_from_excel).pack(side=LEFT, padx=5)

    def save_combobox():
        name = name_var.get().strip()
        if not name:
            messagebox.showwarning("Ошибка", "Введите имя списка.", parent=list_window)
            return
        all_configs = (load_json(path, '') for path in
                       [FIELDS_CONFIG_PATH, COMBOBOX_REGULAR_PATH, COMBOBOX_MAINKEY_PATH, COMBINATION_CONFIG_PATH])
        all_names = {item['name'] for config in all_configs for item in config}
        if name != old_name and name in all_names:
            messagebox.showwarning("Ошибка", f"Имя '{name}' уже используется.", parent=list_window)
            return

        if not main_key_var.get():
            widget = sets_list[0].get('widget_ref')
            if not widget: return
            values = [v.strip() for v in widget.get("1.0", "end-1c").split('\n') if v.strip()]
            combo_data = {"name": name, "type": "текст", "tag_type": "комбобокс", "values": sorted(values)}
            config_path, other_config_path = COMBOBOX_REGULAR_PATH, COMBOBOX_MAINKEY_PATH
        else:
            main_keys = []
            for s in sets_list:
                mk = s["main_key"].get().strip()
                if not mk: continue
                kv_pairs = {kv["key"].get().strip(): kv["value"].get().strip()
                            for kv in s["key_values"] if kv["key"].get().strip()}
                if kv_pairs:
                    main_keys.append({mk: kv_pairs})
            combo_data = {"name": name, "type": "текст", "tag_type": "список", "main_keys": main_keys}
            config_path, other_config_path = COMBOBOX_MAINKEY_PATH, COMBOBOX_REGULAR_PATH

        if is_edit:
            original_file_path = COMBOBOX_REGULAR_PATH if original_tag_type == 'комбобокс' else COMBOBOX_MAINKEY_PATH
            cfg = load_json(original_file_path, '')
            cfg = [item for item in cfg if item.get('name') != old_name]
            save_json(original_file_path, cfg)
            if config_path != original_file_path:
                other_cfg = load_json(other_config_path, '')
                other_cfg = [item for item in other_cfg if item.get('name') != old_name]
                save_json(other_config_path, other_cfg)

        final_config = load_json(config_path, '')
        final_config.append(combo_data)
        save_json(config_path, final_config)
        mapping = {}

        if original_tag_type == 'список' and is_edit:
            # Load the old version from disk to compare
            old_cfg = load_json(COMBOBOX_MAINKEY_PATH, '')
            old_item = next((x for x in old_cfg if x['name'] == old_name), None)
            if old_item:
                old_subkeys = []
                for mk in old_item.get('main_keys', []):
                    old_subkeys.extend(list(list(mk.values())[0].keys()))

                new_subkeys = []
                for mk in combo_data.get('main_keys', []):
                    new_subkeys.extend(list(list(mk.values())[0].keys()))

                removed = [s for s in old_subkeys if s not in new_subkeys]
                added = [s for s in new_subkeys if s not in old_subkeys]
                if len(removed) == len(added):
                    mapping.update(dict(zip(removed, added)))

        # Always include the main tag rename if changed
        if is_edit and old_name != name:
            mapping[old_name] = name

        if mapping:
            update_references(mapping)

        refresh_all_windows(listbox)
        list_window.destroy()

    def import_from_excel():
        """
        Handles import. If Main Key is NOT checked, runs batch import for ALL regular comboboxes
        from import_reg.xlsx. If Main Key IS checked, runs import for the specific main key list
        from import.xlsx.
        """
        global window

        # --- NEW LOGIC FOR REGULAR COMBOBOX (Main Key NOT selected) ---
        if not main_key_var.get():
            xlsx_path = os.path.join(IMPORT_FLD, "import_reg.xlsx")
            xls_path = os.path.join(IMPORT_FLD, "import_reg.xls")

            if os.path.exists(xlsx_path):
                file_path = xlsx_path
            elif os.path.exists(xls_path):
                file_path = xls_path
            else:
                messagebox.showwarning("Ошибка", f"Файл 'import_reg.xlsx' (или .xls) не найден в папке import_fld",
                                       parent=list_window)
                return

            # 1. Confirmation
            if not messagebox.askyesno("Подтверждение импорта (Списки)",
                                       f"Обнаружен файл '{os.path.basename(file_path)}'.\n"
                                       "Запустить БАТЧ-импорт: все комбобоксы и их значения будут созданы/обновлены в конфигурации.\n"
                                       "Продолжить?",
                                       parent=list_window):
                return

            try:
                # 2. Load Excel file and parse data
                wb = openpyxl.load_workbook(file_path)
                sheet = wb.active

                if sheet.max_column < 2:
                    messagebox.showerror("Ошибка",
                                         "Ошибка: Файл 'import_reg' должен содержать 2 столбца (имя, значение).",
                                         parent=list_window)
                    return

                combos_to_import = defaultdict(set)
                rows_processed = 0
                current_name = ""
                for row in sheet.iter_rows(min_row=1, values_only=True):
                    combo_name_val = str(row[0]).strip() if row[0] is not None else ""
                    combo_value = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""

                    if combo_name_val:
                        current_name = combo_name_val

                    if current_name and combo_value:
                        combos_to_import[current_name].add(combo_value)
                        rows_processed += 1

                if not combos_to_import:
                    messagebox.showinfo("Импорт (Списки)", "Не найдено допустимых строк (имя, значение) для импорта.",
                                        parent=list_window)
                    return

                # 3. Load existing config and merge
                combobox_regular_config = load_json(COMBOBOX_REGULAR_PATH, [])
                existing_combos_lookup = {c['name']: c for c in combobox_regular_config}

                created_count = 0
                updated_count = 0
                total_values_added = 0

                # 4. Merge/Create
                for name, new_values_set in combos_to_import.items():
                    if name in existing_combos_lookup:
                        # --- UPDATE EXISTING ---
                        existing_entry = existing_combos_lookup[name]
                        if 'values' not in existing_entry or not isinstance(existing_entry['values'], list):
                            existing_entry['values'] = []

                        existing_values_set = set(existing_entry['values'])
                        newly_added_values = new_values_set - existing_values_set

                        if newly_added_values:
                            existing_values_set.update(newly_added_values)
                            existing_entry['values'] = sorted(list(existing_values_set))
                            updated_count += 1
                            total_values_added += len(newly_added_values)

                    else:
                        # --- CREATE NEW ---
                        new_entry = {
                            "name": name,
                            "type": "текст",
                            "tag_type": "комбобокс",
                            "values": sorted(list(new_values_set))
                        }
                        combobox_regular_config.append(new_entry)
                        created_count += 1
                        total_values_added += len(new_values_set)

                # 5. Save and refresh
                if created_count > 0 or updated_count > 0:
                    save_json(COMBOBOX_REGULAR_PATH, combobox_regular_config)

                # FIX: Use the listbox reference passed to the outer function (constructor_listbox_ref)
                # If your 'open_list_window' function doesn't take this argument, you should change
                # this line to refresh_all_windows() to force a full refresh.
                refresh_all_windows(constructor_listbox_ref)

                messagebox.showinfo(
                    "Импорт завершен",
                    f"Новых списков создано: {created_count}\n"
                    f"Существующих обновлено: {updated_count}\n"
                    f"Всего значений добавлено: {total_values_added}",
                    parent=list_window
                )

                list_window.destroy()  # Close the current list creation window
                return

            except Exception as e:
                messagebox.showerror("Ошибка", f"Произошла непредвиденная ошибка во время импорта: {str(e)}",
                                     parent=list_window)
                print(f"Combobox batch import error: {str(e)}")
                return

        # --- OLD LOGIC FOR MAIN-KEY COMBOBOX (Main Key IS selected) ---
        else:
            import_path = os.path.join(IMPORT_FLD, "import.xlsx")

            if not os.path.exists(import_path):
                messagebox.showwarning("Ошибка", f"Файл {import_path} не найден", parent=list_window)
                return

            try:
                wb = openpyxl.load_workbook(import_path)
                sheet = wb.active
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось открыть файл {import_path}. Ошибка: {str(e)}",
                                     parent=list_window)
                return

            main_keys_data = {}
            current_main_key = None
            for i, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), 1):
                if len(row) < 3 or (row[1] is None or str(row[1]).strip() == "") or \
                        (row[2] is None or str(row[2]).strip() == ""):
                    if any(c is not None for c in row):
                        messagebox.showwarning("Ошибка",
                                               f"Строка {i}: Неверный формат. Ключ и значение должны быть заполнены.",
                                               parent=list_window)
                        return
                    continue
                main_key = str(row[0]).strip() if row[0] is not None else ""
                key = str(row[1]).strip()
                value = str(row[2]).strip()
                if main_key:
                    current_main_key = main_key
                    if current_main_key not in main_keys_data:
                        main_keys_data[current_main_key] = {}
                    main_keys_data[current_main_key][key] = value
                elif current_main_key:
                    if key in main_keys_data[current_main_key]:
                        messagebox.showwarning("Ошибка",
                                               f"Строка {i}: Дублирующийся ключ '{key}' для главного ключа '{current_main_key}'.",
                                               parent=list_window)
                        return
                    main_keys_data[current_main_key][key] = value
                else:
                    messagebox.showwarning("Ошибка",
                                           f"Строка {i}: Ключ '{key}' не имеет предшествующего главного ключа.",
                                           parent=list_window)
                    return
            if not main_keys_data:
                messagebox.showwarning("Ошибка", "В файле не найдено данных для импорта.", parent=list_window)
                return
            total_sub_keys = sum(len(d) for d in main_keys_data.values())
            if messagebox.askyesno("Подтверждение импорта",
                                   f"Найдено {len(main_keys_data)} главных ключей и {total_sub_keys} подчинённых ключей.\n\n"
                                   "Это окно будет закрыто, и все текущие данные будут перезаписаны.\n"
                                   "Продолжить?",
                                   parent=list_window):
                sets_list.clear()
                for main_key, sub_dict in main_keys_data.items():
                    new_set = {
                        "main_key": tk.StringVar(value=main_key),
                        "key_values": [{"key": tk.StringVar(value=k), "value": tk.StringVar(value=v)}
                                       for k, v in sub_dict.items()]
                    }
                    sets_list.append(new_set)
                save_combobox()

    # --- Window Layout ---
    top_controls_frame = tk.Frame(list_window)
    top_controls_frame.pack(fill="x", padx=10, pady=5)
    table_frame = tk.Frame(list_window)
    table_frame.pack(padx=10, pady=5, fill="both", expand=True)
    bottom_buttons_frame = tk.Frame(list_window)
    bottom_buttons_frame.pack(pady=10)

    tk.Label(top_controls_frame, text="Имя списка:").grid(row=0, column=0, sticky="w")
    tk.Entry(top_controls_frame, textvariable=name_var, width=40).grid(row=0, column=1, sticky="ew")
    if not is_edit:
        ttk.Checkbutton(top_controls_frame, text="Использовать главный ключ", variable=main_key_var,
                        command=refresh_table).grid(row=1, column=0, columnspan=2, pady=5)
    top_controls_frame.grid_columnconfigure(1, weight=1)

    tk.Button(bottom_buttons_frame, text="ОК", width=10, command=save_combobox).pack(side=LEFT, padx=5)
    tk.Button(bottom_buttons_frame, text="Отмена", width=10, command=list_window.destroy).pack(side=LEFT, padx=5)

    refresh_table()


# --- Combination Window Functions ---
def open_combination_window(listbox, item_to_edit, parent_window):
    """
    Window to create or edit a 'Combination' tag.
    If item_to_edit is provided, it opens in edit mode.
    """
    is_edit = item_to_edit is not None
    title = "Редактирование сочетания" if is_edit else "Создание сочетания"

    combo_window = tk.Toplevel(parent_window)
    combo_window.title(title)
    combo_window.resizable(False, False)
    combo_window.focus_set()
    combo_window.grab_set()

    name_var = tk.StringVar()
    combination_tags = []
    old_name = None

    if is_edit:
        old_name, _, _ = listbox.item(item_to_edit)['values']
        combo_config = load_json(COMBINATION_CONFIG_PATH, 'combination_config')
        existing_combo_data = next((item for item in combo_config if item['name'] == old_name), None)
        if existing_combo_data:
            name_var.set(existing_combo_data['name'])
            combination_tags.extend(existing_combo_data['tags'])
        else:
            messagebox.showwarning("Ошибка", "Данные сочетания не найдены.")
            return

    all_tags = get_all_tags_for_constructor()
    numbers = load_number_config()
    for num in numbers:
        all_tags.append(num['name'])
    sorted_tags = sorted(list(all_tags))

    def refresh_listbox():
        for item in combo_listbox.get_children():
            combo_listbox.delete(item)
        for i, tag in enumerate(combination_tags):
            display_tag = tag
            if tag == ' ':
                display_tag = "[Пробел]"
            elif tag == '\n':
                display_tag = "[Абзац]"
            elif tag == 'today_tag':
                display_tag = "[СЕГОДНЯ]"
            combo_listbox.insert("", "end", values=(i + 1, display_tag))

    def add_element(element):
        if element:
            if element == "[Пробел]":
                element = ' '
            elif element == "[Абзац]":
                element = '\n'
            elif element == "[СЕГОДНЯ]":
                element = 'today_tag'

            combination_tags.append(element)
            refresh_listbox()

    def remove_element():
        selected_item = combo_listbox.selection()
        if selected_item:
            index_to_remove = int(combo_listbox.item(selected_item[0])['values'][0]) - 1
            if 0 <= index_to_remove < len(combination_tags):
                del combination_tags[index_to_remove]
                refresh_listbox()

    def move_element(direction):
        selected_item = combo_listbox.selection()
        if not selected_item:
            return
        index = int(combo_listbox.item(selected_item[0])['values'][0]) - 1
        if direction == 'up' and index > 0:
            combination_tags[index], combination_tags[index - 1] = combination_tags[index - 1], combination_tags[index]
        elif direction == 'down' and index < len(combination_tags) - 1:
            combination_tags[index], combination_tags[index + 1] = combination_tags[index + 1], combination_tags[index]
        refresh_listbox()
        new_selection_index = index + (1 if direction == 'down' else -1)
        combo_listbox.selection_set(combo_listbox.get_children()[new_selection_index])

    def save_combination():
        name = name_var.get().strip()
        if not name:
            messagebox.showwarning("Ошибка", "Введите имя для сочетания.", parent=combo_window)
            return
        if not combination_tags:
            messagebox.showwarning("Ошибка", "Добавьте хотя бы один элемент в сочетание.", parent=combo_window)
            return

        all_configs = (load_json(path, '') for path in
                       [FIELDS_CONFIG_PATH, COMBOBOX_REGULAR_PATH, COMBOBOX_MAINKEY_PATH, COMBINATION_CONFIG_PATH])
        all_names = {item['name'] for config in all_configs for item in config if item['name'] != old_name}
        if name in all_names:
            messagebox.showwarning("Ошибка", f"Имя '{name}' уже используется.", parent=combo_window)
            return

        combo_data = {"name": name, "type": "текст", "tag_type": "сочетание", "tags": combination_tags}
        combo_config = load_json(COMBINATION_CONFIG_PATH, 'combination_config')

        if is_edit:
            found = False
            for i, combo in enumerate(combo_config):
                if combo['name'] == old_name:
                    combo_config[i] = combo_data
                    found = True
                    break
            if not found:
                messagebox.showerror("Ошибка", "Не удалось найти сочетание для редактирования.")
                return
        else:
            combo_config.append(combo_data)

        save_json(COMBINATION_CONFIG_PATH, combo_config)
        refresh_all_windows(listbox)
        combo_window.destroy()

    content_frame = tk.Frame(combo_window, padx=10, pady=10)
    content_frame.pack(fill="both", expand=True)

    button_frame = tk.Frame(combo_window)
    button_frame.pack(side="bottom", pady=10)
    tk.Button(button_frame, text="ОК", width=10, command=save_combination).pack(side="left", padx=5)
    tk.Button(button_frame, text="Отмена", width=10, command=combo_window.destroy).pack(side="left", padx=5)

    tk.Label(content_frame, text="Имя сочетания:").pack(anchor="w")
    name_entry = tk.Entry(content_frame, textvariable=name_var, width=50)
    name_entry.pack(pady=(0, 10), fill="x")

    options_frame = tk.Frame(content_frame)
    options_frame.pack(fill="x")

    tag_options_frame = tk.Frame(options_frame)
    tag_options_frame.pack(side="left", padx=(0, 10))
    tk.Label(tag_options_frame, text="Добавить тег:").pack(anchor="w")
    tag_combo = ttk.Combobox(tag_options_frame, values=sorted_tags, state="readonly")
    tag_combo.pack(side="left", fill="x", expand=True, padx=(0, 5))
    tk.Button(tag_options_frame, text="Добавить", command=lambda: add_element(tag_combo.get())).pack(side="left")

    literal_options_frame = tk.Frame(options_frame)
    literal_options_frame.pack(side="left")
    tk.Label(literal_options_frame, text="Добавить текст:").pack(anchor="w")
    literal_entry = tk.Entry(literal_options_frame, width=20)
    literal_entry.pack(side="left", fill="x", expand=True, padx=(0, 5))
    tk.Button(literal_options_frame, text="Добавить", command=lambda: add_element(literal_entry.get())).pack(
        side="left")

    literal_buttons_frame = tk.Frame(content_frame, pady=5)
    literal_buttons_frame.pack(fill="x")
    tk.Button(literal_buttons_frame, text="Пробел", command=lambda: add_element(' ')).pack(side="left", padx=2)
    tk.Button(literal_buttons_frame, text="Абзац", command=lambda: add_element('\n')).pack(side="left", padx=2)
    tk.Button(literal_buttons_frame, text="СЕГОДНЯ", command=lambda: add_element('today_tag')).pack(side="left", padx=2)

    tk.Label(content_frame, text="Элементы сочетания:").pack(anchor="w", pady=(10, 0))

    combo_listbox_frame = tk.Frame(content_frame)
    combo_listbox_frame.pack(fill="both", expand=True, pady=(0, 10))

    combo_listbox = ttk.Treeview(combo_listbox_frame, columns=("№", "Элемент"), show="headings", height=10)
    combo_listbox.heading("№", text="№", anchor="center")
    combo_listbox.heading("Элемент", text="Элемент")
    combo_listbox.column("№", width=30, anchor="center")
    combo_listbox.column("Элемент", width=400, anchor="w")

    listbox_scrollbar = ttk.Scrollbar(combo_listbox_frame, orient="vertical", command=combo_listbox.yview)
    combo_listbox.configure(yscrollcommand=listbox_scrollbar.set)
    combo_listbox.pack(side="left", fill="both", expand=True)
    listbox_scrollbar.pack(side="right", fill="y")

    actions_frame = tk.Frame(content_frame)
    actions_frame.pack(fill="x", pady=5)
    tk.Button(actions_frame, text="Удалить", width=12, command=remove_element).pack(side="left", padx=2)
    tk.Button(actions_frame, text="Вверх", width=12, command=lambda: move_element('up')).pack(side="left", padx=2)
    tk.Button(actions_frame, text="Вниз", width=12, command=lambda: move_element('down')).pack(side="left", padx=2)

    refresh_listbox()

def open_number_window(listbox, item_to_edit, parent_window):
    """Window to create/edit Число tag (numeric formula)."""
    is_edit = item_to_edit is not None
    title = "Редактировать число" if is_edit else "Создание числа"

    num_window = tk.Toplevel(parent_window)
    num_window.title(title)
    num_window.resizable(False, False)
    num_window.focus_set()
    num_window.grab_set()

    name_var = tk.StringVar()
    sequence = []
    old_name = None

    if is_edit:
        old_name, _, _ = listbox.item(item_to_edit)['values']
        number_config = load_number_config()
        existing = next((x for x in number_config if x['name'] == old_name), None)
        if existing:
            name_var.set(existing['name'])
            sequence = existing['sequence'][:]

    all_tags = list(get_all_tags_for_constructor())
    all_tags = sorted(all_tags)
    operators = ["+", "-", "*", "/"]

    def refresh_listbox():
        num_listbox.delete(*num_listbox.get_children())
        for i, elem in enumerate(sequence):
            num_listbox.insert("", "end", values=(i + 1, elem))

    def add_element(elem):
        if elem:
            sequence.append(elem)
            refresh_listbox()

    def remove_element():
        sel = num_listbox.selection()
        if sel:
            idx = int(num_listbox.item(sel[0])['values'][0]) - 1
            if 0 <= idx < len(sequence):
                del sequence[idx]
                refresh_listbox()

    def move_element(direction):
        sel = num_listbox.selection()
        if not sel:
            return
        idx = int(num_listbox.item(sel[0])['values'][0]) - 1
        if direction == "up" and idx > 0:
            sequence[idx], sequence[idx-1] = sequence[idx-1], sequence[idx]
        elif direction == "down" and idx < len(sequence)-1:
            sequence[idx], sequence[idx+1] = sequence[idx+1], sequence[idx]
        refresh_listbox()

    def save_number():
        name = name_var.get().strip()
        if not name:
            messagebox.showwarning("Ошибка", "Введите имя числа.", parent=num_window)
            return
        if not sequence:
            messagebox.showwarning("Ошибка", "Добавьте хотя бы один элемент.", parent=num_window)
            return

        # Check duplicate names
        all_configs = (load_json(path, '') for path in
                       [FIELDS_CONFIG_PATH, COMBOBOX_REGULAR_PATH, COMBOBOX_MAINKEY_PATH,
                        COMBINATION_CONFIG_PATH, NUMBER_CONFIG_PATH])
        all_names = {item['name'] for cfg in all_configs for item in cfg if item['name'] != old_name}
        if name in all_names:
            messagebox.showwarning("Ошибка", f"Имя '{name}' уже используется.", parent=num_window)
            return

        number_config = load_number_config()
        entry = {"name": name, "type": "число", "tag_type": "число", "sequence": sequence}
        if is_edit:
            for i, n in enumerate(number_config):
                if n['name'] == old_name:
                    number_config[i] = entry
                    break
        else:
            number_config.append(entry)

        save_number_config(number_config)
        refresh_all_windows(listbox)
        num_window.destroy()

    # --- UI Layout ---
    top = tk.Frame(num_window)
    top.pack(fill="x", padx=10, pady=5)
    tk.Label(top, text="Имя числа:").pack(side="left")
    tk.Entry(top, textvariable=name_var, width=30).pack(side="left", padx=5)

    mid = tk.Frame(num_window)
    mid.pack(fill="both", expand=True, padx=10, pady=5)

    num_listbox = ttk.Treeview(mid, columns=("№", "Элемент"), show="headings", height=10)
    num_listbox.heading("№", text="№")
    num_listbox.heading("Элемент", text="Элемент")
    num_listbox.column("№", width=40, anchor="center")
    num_listbox.column("Элемент", width=150, anchor="center")
    num_listbox.pack(side="left", fill="both", expand=True)

    btns = tk.Frame(mid)
    btns.pack(side="left", padx=5)
    tk.Button(btns, text="Удалить", command=remove_element).pack(pady=2)
    tk.Button(btns, text="Вверх", command=lambda: move_element("up")).pack(pady=2)
    tk.Button(btns, text="Вниз", command=lambda: move_element("down")).pack(pady=2)

    bottom = tk.Frame(num_window)
    bottom.pack(fill="x", padx=10, pady=5)

    # Tag selector
    tag_combo = ttk.Combobox(bottom, values=all_tags, state="readonly")
    tag_combo.pack(side="left", padx=5)
    tk.Button(bottom, text="Добавить тег", command=lambda: add_element(tag_combo.get())).pack(side="left", padx=5)

    # Digit entry
    digit_var = tk.StringVar()
    digit_entry = tk.Entry(bottom, textvariable=digit_var, width=10)
    digit_entry.pack(side="left", padx=5)
    tk.Button(bottom, text="Добавить число",
              command=lambda: add_element(digit_var.get().replace(".", ",") if digit_var.get() else None)
              ).pack(side="left", padx=5)

    # Operator
    op_combo = ttk.Combobox(bottom, values=operators, state="readonly", width=5)
    op_combo.pack(side="left", padx=5)
    tk.Button(bottom, text="Добавить оператор", command=lambda: add_element(op_combo.get())).pack(side="left", padx=5)

    # Brackets
    tk.Button(bottom, text="(", command=lambda: add_element("(")).pack(side="left", padx=2)
    tk.Button(bottom, text=")", command=lambda: add_element(")")).pack(side="left", padx=2)

    okcancel = tk.Frame(num_window)
    okcancel.pack(pady=10)
    tk.Button(okcancel, text="ОК", width=10, command=save_number).pack(side="left", padx=5)
    tk.Button(okcancel, text="Отмена", width=10, command=num_window.destroy).pack(side="left", padx=5)

    refresh_listbox()


def load_combination_config():
    """Load combination config from JSON file, initializing if not found."""
    try:
        with open(COMBINATION_CONFIG_PATH, encoding='utf8') as f:
            return json.load(f)
    except FileNotFoundError:
        with open(COMBINATION_CONFIG_PATH, 'w', encoding='utf8') as f:
            json.dump([], f, ensure_ascii=False, indent=4)
        messagebox.showwarning("Информация", f"Создан новый файл конфигурации: {COMBINATION_CONFIG_PATH}")
        return []
    except json.JSONDecodeError:
        messagebox.showwarning("Предупреждение",
                               f"Файл {COMBINATION_CONFIG_PATH} поврежден. Инициализация пустой конфигурации.")
        with open(COMBINATION_CONFIG_PATH, 'w', encoding='utf8') as f:
            json.dump([], f, ensure_ascii=False, indent=4)
        return []

def load_combination_config():
    """Load combination config from JSON file, initializing if not found."""
    try:
        with open(COMBINATION_CONFIG_PATH, encoding='utf8') as f:
            return json.load(f)
    except FileNotFoundError:
        with open(COMBINATION_CONFIG_PATH, 'w', encoding='utf8') as f:
            json.dump([], f, ensure_ascii=False, indent=4)
        messagebox.showwarning("Информация", f"Создан новый файл конфигурации: {COMBINATION_CONFIG_PATH}")
        return []
    except json.JSONDecodeError:
        messagebox.showwarning("Предупреждение", f"Файл {COMBINATION_CONFIG_PATH} поврежден. Инициализация пустой конфигурации.")
        with open(COMBINATION_CONFIG_PATH, 'w', encoding='utf8') as f:
            json.dump([], f, ensure_ascii=False, indent=4)
        return []


def open_edit_tag_window(listbox, parent_window):
    """Opens the appropriate window for editing a selected tag."""
    selected_item = listbox.selection()
    if not selected_item:
        messagebox.showwarning("Ошибка", "Выберите тег для редактирования.", parent=listbox)
        return

    item_id = selected_item[0]
    item_values = listbox.item(item_id)['values']

    if not item_values or len(item_values) < 3:
        messagebox.showwarning("Ошибка", "Некорректные данные тега.", parent=listbox)
        return

    tag_type = item_values[2]

    # Pass the listbox and the item_id to the specific edit window
    if tag_type == 'поле':
        open_field_window(listbox, item_id, parent_window)
    elif tag_type == 'чекбокс':
        open_checkbox_window(listbox, item_id, parent_window)
    elif tag_type == 'комбобокс' or tag_type == 'список':
        open_list_window(listbox, item_id, parent_window)
    elif tag_type == 'сочетание':
        open_combination_window(listbox, item_id, parent_window)
    elif tag_type == "число":
        open_number_window(listbox, item_id, parent_window)
    else:
        messagebox.showinfo("Информация", "Редактирование для этого типа тега еще не реализовано.")


def delete_tag(tags_listbox, parent_window):
    """
    Deletes one or more selected tags and performs a cascading delete
    of their subkeys from rules, combinations, and numbers.
    """
    selected_items = tags_listbox.selection()
    if not selected_items:
        messagebox.showwarning("Предупреждение", "Пожалуйста, выберите тег(и) для удаления.", parent=parent_window)
        return

    # Collect details of all selected tags for the confirmation message
    tags_to_delete = []
    for item_id in selected_items:
        item_values = tags_listbox.item(item_id)['values']
        tag_name = item_values[0]
        tag_type = item_values[2]
        tags_to_delete.append({'name': tag_name, 'type': tag_type})

    # Format a user-friendly confirmation message, now with a warning
    names_str = "\n- ".join([t['name'] for t in tags_to_delete])
    if not messagebox.askyesno(
        "Подтверждение",
        f"Вы уверены, что хотите удалить следующие теги?\n\n- {names_str}\n\n"
        "ВНИМАНИЕ: Все выбранные теги и их дочерние элементы (для списков) будут также удалены "
        "из всех правил, сочетаний и чисел.",
        parent=parent_window
    ):
        return

    try:
        # --- START: New Cascading Delete Logic ---

        # 1. Collect all top-level tags to purge
        tags_to_purge = set(tag['name'] for tag in tags_to_delete)

        # 2. Find all subkeys from any 'список' type tags being deleted.
        subkeys_to_purge = set()
        list_tags_to_delete = [tag['name'] for tag in tags_to_delete if tag['type'] == 'список']

        if list_tags_to_delete:
            mainkey_config = load_json(COMBOBOX_MAINKEY_PATH, 'combobox_mainkey')
            for combo in mainkey_config:
                if combo.get('name') in list_tags_to_delete:
                    for mk_dict in combo.get('main_keys', []):
                        if mk_dict and isinstance(list(mk_dict.values())[0], dict):
                            subkeys_dict = list(mk_dict.values())[0]
                            for subkey_name in subkeys_dict.keys():
                                subkeys_to_purge.add(subkey_name)

        # 3. Add subkeys to the purge set
        tags_to_purge.update(subkeys_to_purge)

        # 4. If we have tags to purge, clean them up from other configs.
        if tags_to_purge:
            # 4a. Clean up Combination Config
            combination_config = load_json(COMBINATION_CONFIG_PATH, 'combination_config')
            for combo in combination_config:
                combo['sequence'] = [elem for elem in combo.get('sequence', []) if elem not in tags_to_purge]
            save_json(COMBINATION_CONFIG_PATH, combination_config)

            # 4b. Clean up Rules Config
            rules_config = load_json(RULES_CONFIG_PATH, 'rules_config')
            for rule in rules_config:
                rule['conditions'] = [cond for cond in rule.get('conditions', []) if cond.get('tag') not in tags_to_purge]
                rule['behaviors'] = [beh for beh in rule.get('behaviors', []) if beh.get('tag') not in tags_to_purge]
            save_json(RULES_CONFIG_PATH, rules_config)

            # 4c. Clean up Number Config
            number_config = load_json(NUMBER_CONFIG_PATH, 'number_config')
            for num in number_config:
                num['sequence'] = [elem for elem in num.get('sequence', []) if elem not in tags_to_purge]
            save_json(NUMBER_CONFIG_PATH, number_config)

        # --- END: New Cascading Delete Logic ---

        # --- Original Deletion Logic (for the main tags themselves) ---
        deletions_by_file = {
            FIELDS_CONFIG_PATH: set(),
            COMBOBOX_REGULAR_PATH: set(),
            COMBOBOX_MAINKEY_PATH: set(),
            COMBINATION_CONFIG_PATH: set(),
            NUMBER_CONFIG_PATH: set()
        }

        for tag in tags_to_delete:
            tag_name = tag['name']
            tag_type = tag['type']
            if tag_type in ('поле', 'чекбокс'):
                deletions_by_file[FIELDS_CONFIG_PATH].add(tag_name)
            elif tag_type == 'комбобокс':
                deletions_by_file[COMBOBOX_REGULAR_PATH].add(tag_name)
            elif tag_type == 'список':
                deletions_by_file[COMBOBOX_MAINKEY_PATH].add(tag_name)
            elif tag_type == 'сочетание':
                deletions_by_file[COMBINATION_CONFIG_PATH].add(tag_name)
            elif tag_type == 'число':
                deletions_by_file[NUMBER_CONFIG_PATH].add(tag_name)

        # Process each configuration file that has items to be deleted
        for config_path, names_to_delete_set in deletions_by_file.items():
            if not names_to_delete_set:
                continue

            config_data = load_json(config_path, '')
            updated_data = [item for item in config_data if item.get('name') not in names_to_delete_set]
            save_json(config_path, updated_data)

        # Refresh the UI once after all deletions are complete
        refresh_all_windows(tags_listbox)
        messagebox.showinfo("Успех", "Выбранные теги и все их дочерние элементы были успешно удалены.", parent=parent_window)

    except Exception as e:
        messagebox.showerror("Ошибка", f"Произошла ошибка при удалении тегов: {e}", parent=parent_window)



# This function is needed to refresh the constructor listbox from other windows
def populate_tags_listbox_in_constructor(listbox):
    for i in listbox.get_children():
        listbox.delete(i)

    all_items = []
    all_items.extend(
        [(f['name'], f['type'], f.get('tag_type', 'поле')) for f in load_json(FIELDS_CONFIG_PATH, 'fields_config')])
    all_items.extend([(c['name'], c['type'], c.get('tag_type', 'комбобокс')) for c in
                      load_json(COMBOBOX_REGULAR_PATH, 'combobox_regular')])
    all_items.extend([(c['name'], c['type'], c.get('tag_type', 'список')) for c in
                      load_json(COMBOBOX_MAINKEY_PATH, 'combobox_mainkey')])
    all_items.extend([(c['name'], c['type'], c.get('tag_type', 'сочетание')) for c in
                      load_json(COMBINATION_CONFIG_PATH, 'combination_config')])

    all_items.sort(key=lambda x: x[0])
    for item in all_items:
        listbox.insert("", "end", values=item)


def refresh_main_and_constructor():
    """Refreshes the dynamic widgets in the main window and the tags in the constructor window."""
    global dynamic_frame

    # 1. Save current state before destroying widgets
    current_state = get_current_widget_values(dynamic_frame)

    if dynamic_frame and dynamic_frame.winfo_exists():
        for widget in dynamic_frame.winfo_children():
            widget.destroy()
        # 2. Pass the saved state to the loading function
        load_all_dynamic_widgets(initial_state=current_state)

    # Find and refresh the constructor's tags listbox if it is open
    for child in window.winfo_children():
        if isinstance(child, tk.Toplevel) and child.title() == "Конструктор":
            for grand_child in child.winfo_children():
                if isinstance(grand_child, ttk.Notebook):
                    tags_tab = grand_child.winfo_children()[0]
                    for great_grand_child in tags_tab.winfo_children():
                        if isinstance(great_grand_child, tk.Frame):
                            for great_great_grand_child in great_grand_child.winfo_children():
                                if isinstance(great_great_grand_child, ttk.Treeview):
                                    populate_tags_listbox(great_great_grand_child)
                                    return

def evaluate_condition(condition, merge_data):
    """
    Evaluate a single condition against the main window's tag value.
    Args:
        condition (dict) {'tag': 'tag_name', 'condition': 'содержит', 'rule': 'value'}
        merge_data (dict): Data from get_common_merge_data()
    Returns:
        bool: True if condition is met, False otherwise
    """
    tag = condition['tag']
    cond_type = condition['condition']
    rule = condition['rule']

    if tag not in merge_data:
        return False  # Skip missing tags silently

    value = merge_data[tag]
    try:
        if cond_type == 'содержит':
            return rule.lower() in str(value).lower()
        elif cond_type == 'начинается с':
            return str(value).lower().startswith(rule.lower())
        elif cond_type == 'заканчивается на':
            return str(value).lower().endswith(rule.lower())
        elif cond_type == 'больше':
            # MODIFIED: Replace comma with dot for float conversion
            return float(str(value).replace(',', '.')) > float(str(rule).replace(',', '.'))
        elif cond_type == 'меньше':
            # MODIFIED: Replace comma with dot for float conversion
            return float(str(value).replace(',', '.')) < float(str(rule).replace(',', '.'))
        elif cond_type == 'равно':
            if tag in checkbox_vars:
                return value == rule
            else:
                # MODIFIED: Replace comma with dot for float conversion
                return float(str(value).replace(',', '.')) == float(str(rule).replace(',', '.'))
        elif cond_type == 'True':
            return value == "1"  # Checkbox check
        elif cond_type == 'False':
            return value == "0"  # Checkbox check
        else:
            return False  # Unknown condition
    except (ValueError, AttributeError):
        # Catch errors from float conversion on non-numeric text or attribute errors on non-strings
        return False

def apply_behaviors(behaviors, merge_data):
    """
    Apply behaviors to merge_data.
    Args:
        behaviors (list): List of behavior dictionaries from rules_config.json
        merge_data (dict): Data from get_common_merge_data()
    Returns:
        dict: Modified merge_data
    """
    for behavior in behaviors:
        tag = behavior['tag']
        action = behavior['condition']
        rule = behavior['rule']
        if tag not in merge_data:
            continue
        value = str(merge_data[tag])  # Convert to string for safety
        try:
            if action == "очистить":
                merge_data[tag] = ""
            elif action == "очистить при не выполнении":
                merge_data[tag] = ""
            elif action == "CAPS":
                merge_data[tag] = value.upper()
            elif action == "верхняя буква":
                merge_data[tag] = value.capitalize()
            elif action == "нижняя буква":
                merge_data[tag] = value.lower()
            elif action == "транслит":
                merge_data[tag] = translit(value, rule)
            elif action == "добавить текст в начале":
                merge_data[tag] = rule + str(value)
            elif action == "добавить текст в конце":
                merge_data[tag] = str(value) + rule
            elif action == "добавить дней кален":
                date_obg = datetime.strptime(value, '%d.%m.%Y')
                days = int(rule)
                merge_data[tag] = (date_obg + timedelta(days=days)).strftime('%d.%m.%Y')
            elif action == "отнять дней кален":
                date_obg = datetime.strptime(value, '%d.%m.%Y')
                days = int(rule)
                merge_data[tag] = (date_obg - timedelta(days=days)).strftime('%d.%m.%Y')
            elif action == "добавить рабочих дней":
                date_obg = datetime.strptime(value, '%d.%m.%Y')
                days = int(rule)
                np_date = np.datetime64(date_obg, 'D')
                result_date = np.busday_offset(np_date, days, roll='forward')
                merge_data[tag] = result_date.astype('datetime64[ms]').astype(datetime).strftime('%d.%m.%Y')
            elif action == "отнять рабочих дней":
                date_obg = datetime.strptime(value, '%d.%m.%Y')
                days = int(rule)
                np_date = np.datetime64(date_obg, 'D')
                result_date = np.busday_offset(np_date, -days, roll='forward')
                merge_data[tag] = result_date.astype('datetime64[ms]').astype(datetime).strftime('%d.%m.%Y')
            elif action == "обрезать":
                if rule == "":
                    raise ValueError("Rule cannot be empty")

                # Handle single number (e.g., ":4" or "4:")
                if rule.startswith(":"):
                    try:
                        end = int(rule[1:])
                        if end >= 0:
                            merge_data[tag] = value[:-end] if end <= len(value) else value
                        else:
                            raise ValueError("End index must be non-negative")
                    except ValueError:
                        raise ValueError("Rule must be in ':end' format with a valid integer")
                elif rule.endswith(":"):
                    try:
                        start = int(rule[:-1])
                        if start >= 0:
                            merge_data[tag] = value[start:] if start < len(value) else value
                        else:
                            raise ValueError("Start index must be non-negative")
                    except ValueError:
                        raise ValueError("Rule must be in 'start:' format with a valid integer")
                else:
                    # Existing logic for "start:end" format
                    if ':' not in rule:
                        raise ValueError("Rule must be in 'start:end', ':end', or 'start:' format")
                    start, end = map(int, rule.split(':'))
                    if start <= end:
                        merge_data[tag] = value[:start] + value[end + 1:] if 0 <= start <= end < len(value) else value
                    else:
                        merge_data[tag] = value[:end] + value[start + 1:] if 0 <= end <= start < len(value) else value

            # --- NEW: числа в слова ---
            elif action == "числа в слова":
                try:
                    # rule is encoded as "lang|currency"
                    parts = (rule or "").split("|")
                    lang = parts[0] if len(parts) > 0 and parts[0] else "uk"
                    currency = parts[1] if len(parts) > 1 and parts[1] else ""

                    raw_val = value.replace(",", ".")
                    num = float(raw_val) if raw_val.strip() else 0.0

                    if currency:
                        merge_data[tag] = num2words(num, lang=lang, to="currency", currency=currency)
                    else:
                        merge_data[tag] = num2words(num, lang=lang)  # always cardinal
                except Exception as ne:
                    print(f"[числа в слова] error for {tag}: {ne}")

        except Exception as e:
            print(f"Error applying behavior {action} on {tag}: {e}")
    return merge_data


getcontext().prec = 28

def evaluate_number_sequence(sequence, merge_data):
    expr_parts = []
    for elem in sequence:
        # --- tag from merge_data ---
        if elem in merge_data:
            val = merge_data.get(elem, "0")
            try:
                val = str(val).replace(",", ".")
                Decimal(val)  # check if numeric
            except:
                val = "0"
            expr_parts.append(f"Decimal('{val}')")
            continue

        # --- subkey from main_key selections ---
        found_subkey = False
        for subdict in main_key_selections.values():
            if elem in subdict:
                val = subdict.get(elem, "0")
                try:
                    val = str(val).replace(",", ".")
                    Decimal(val)
                except:
                    val = "0"
                expr_parts.append(f"Decimal('{val}')")
                found_subkey = True
                break
        if found_subkey:
            continue

        # --- math operators ---
        if elem in ["+", "-", "*", "/", "(", ")"]:
            expr_parts.append(elem)
            continue

        # --- numeric literal ---
        try:
            val = str(elem).replace(",", ".")
            Decimal(val)
        except:
            val = "0"
        expr_parts.append(f"Decimal('{val}')")

    expr = " ".join(expr_parts)
    try:
        result = eval(expr, {"Decimal": Decimal})
        formatted = f"{result:.2f}".replace(".", ",")
        return formatted
    except Exception:
        return "0,00"


def open_create_rule_window(listbox, constructor_window):
    create_rule_window = tk.Toplevel(constructor_window)
    create_rule_window.title("Создать правило")
    create_rule_window.geometry("1070x230")
    create_rule_window.resizable(False, False)
    create_rule_window.focus_set()
    create_rule_window.grab_set()

    # Get all tags, sort them, and add a blank option at the start
    all_tags = list(set(get_all_tags_for_constructor()))

    sorted_tags = sorted(all_tags)
    sorted_tags.insert(0, '')

    main_frame = tk.Frame(create_rule_window)
    main_frame.pack(padx=10, pady=5, fill=tk.BOTH, expand=True)

    widgets_3 = []
    widgets_4 = []

    tk.Label(main_frame, text="Имя правила:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
    name_var = tk.StringVar()
    tk.Entry(main_frame, textvariable=name_var, width=30).grid(row=0, column=1, columnspan=5, padx=5, pady=5,
                                                               sticky="w")

    # --- LEFT COLUMN (CONDITIONS) ---
    tk.Label(main_frame, text="Теги", anchor="center").grid(row=1, column=0, padx=5, pady=5)
    tk.Label(main_frame, text="Условие", anchor="center").grid(row=1, column=1, padx=5, pady=5)
    tk.Label(main_frame, text="Правило", anchor="center").grid(row=1, column=2, padx=5, pady=5)

    first_column_frame = tk.Frame(main_frame)
    first_column_frame.grid(row=2, column=0, columnspan=3, padx=5, pady=5, sticky="w")
    tag_var1 = tk.StringVar()
    tag_combobox1 = ttk.Combobox(first_column_frame, textvariable=tag_var1, width=20, state="readonly",
                                 values=sorted_tags)
    tag_combobox1.grid(row=0, column=0, padx=5, pady=5)
    condition_var1 = tk.StringVar()
    condition_combobox1 = ttk.Combobox(first_column_frame, textvariable=condition_var1, width=20, state="readonly",
                                       values=["", "содержит", "начинается с", "заканчивается на", "меньше", "больше",
                                               "равно", "True", "False"])
    condition_combobox1.grid(row=0, column=1, padx=5, pady=5)
    rule_entry_var1 = tk.StringVar()
    rule_entry1 = tk.Entry(first_column_frame, textvariable=rule_entry_var1, width=20)
    rule_entry1.grid(row=0, column=2, padx=5, pady=5)

    def toggle_rule_entry_state(condition_combobox, rule_entry):
        selected_condition = condition_combobox.get()
        rule_entry.config(state="disabled" if selected_condition in ["True", "False"] else "normal")

    condition_combobox1.bind("<<ComboboxSelected>>",
                             lambda e: toggle_rule_entry_state(condition_combobox1, rule_entry1))
    widgets_3.append((first_column_frame, tag_combobox1, condition_combobox1, rule_entry1, rule_entry_var1))

    def add_new_3_item_set():
        base_row = len(widgets_3)
        new_frame = tk.Frame(main_frame)
        new_frame.grid(row=2 + base_row, column=0, columnspan=3, padx=5, pady=5, sticky="w")
        new_tag_var = tk.StringVar()
        new_tag_combobox = ttk.Combobox(new_frame, textvariable=new_tag_var, width=20, state="readonly",
                                        values=sorted_tags)
        new_tag_combobox.grid(row=0, column=0, padx=5, pady=5)
        new_condition_var = tk.StringVar()
        new_condition_combobox = ttk.Combobox(new_frame, textvariable=new_condition_var, width=20, state="readonly",
                                              values=["", "содержит", "начинается с", "заканчивается на", "меньше",
                                                      "больше", "равно", "True", "False"])
        new_condition_combobox.grid(row=0, column=1, padx=5, pady=5)
        new_rule_var = tk.StringVar()
        new_rule_entry = tk.Entry(new_frame, textvariable=new_rule_var, width=20)
        new_rule_entry.grid(row=0, column=2, padx=5, pady=5)
        new_condition_combobox.bind("<<ComboboxSelected>>",
                                    lambda e: toggle_rule_entry_state(new_condition_combobox, new_rule_entry))
        widgets_3.append((new_frame, new_tag_combobox, new_condition_combobox, new_rule_entry, new_rule_var))
        buttons_frame_3.grid(row=3 + base_row, column=0, columnspan=3, padx=5, pady=5, sticky="w")
        new_height = max(230, 230 + max(len(widgets_3), len(widgets_4)) * 40)
        create_rule_window.geometry(f"1070x{new_height}")

    def remove_last_3_item_set():
        if len(widgets_3) > 1:
            frame_to_destroy, _, _, _, _ = widgets_3.pop()
            frame_to_destroy.destroy()
            buttons_frame_3.grid(row=2 + len(widgets_3), column=0, columnspan=3, padx=5, pady=5, sticky="w")
            new_height = max(230, 230 + max(len(widgets_3), len(widgets_4)) * 40)
            create_rule_window.geometry(f"1070x{new_height}")

    buttons_frame_3 = tk.Frame(main_frame)
    buttons_frame_3.grid(row=2 + len(widgets_3), column=0, columnspan=3, padx=5, pady=5, sticky="w")
    ttk.Button(buttons_frame_3, text="Добавить", command=add_new_3_item_set).pack(side="left", padx=(0, 5))
    ttk.Button(buttons_frame_3, text="Удалить", command=remove_last_3_item_set).pack(side="left")

    ttk.Separator(main_frame, orient='vertical').grid(row=1, column=3, rowspan=100, sticky="ns", padx=10)

    # --- RIGHT COLUMN (BEHAVIORS) ---
    tk.Label(main_frame, text="Теги", anchor="center").grid(row=1, column=4, padx=5, pady=5)
    tk.Label(main_frame, text="Условие", anchor="center").grid(row=1, column=5, padx=5, pady=5)
    tk.Label(main_frame, text="Правило", anchor="center").grid(row=1, column=6, padx=5, pady=5)
    tk.Label(main_frame, text="Опция", anchor="center").grid(row=1, column=7, padx=5, pady=5)

    second_column_frame = tk.Frame(main_frame)
    second_column_frame.grid(row=2, column=4, columnspan=4, padx=5, pady=5, sticky="w")
    tag_var2 = tk.StringVar()
    tag_combobox2 = ttk.Combobox(second_column_frame, textvariable=tag_var2, width=20, state="readonly",
                                 values=sorted_tags)
    tag_combobox2.grid(row=0, column=0, padx=5, pady=5)
    condition_var2 = tk.StringVar()
    condition_combobox2 = ttk.Combobox(second_column_frame, textvariable=condition_var2, width=20, state="readonly",
                                       values=["", "очистить", "очистить при не выполнении", "CAPS", "верхняя буква",
                                               "нижняя буква", "транслит", "числа в слова", "добавить текст в начале",
                                               "добавить текст в конце", "добавить дней кален", "отнять дней кален", "добавить рабочих дней", "отнять рабочих дней", "обрезать"])
    condition_combobox2.grid(row=0, column=1, padx=5, pady=5)
    rule_entry_var2 = tk.StringVar()
    rule_entry2 = tk.Entry(second_column_frame, textvariable=rule_entry_var2, width=20)  # Placeholder
    option_entry_var2 = tk.StringVar()
    option_entry2 = tk.Entry(second_column_frame, textvariable=option_entry_var2, width=20, state="disabled")
    option_entry2.grid(row=0, column=3, padx=5, pady=5)

    # <<< FIX 1: Store the rule's StringVar (rule_entry_var2) in the widgets list for the first row
    widgets_4.append(
        (second_column_frame, tag_combobox2, condition_combobox2, rule_entry2, option_entry2, rule_entry_var2))

    def toggle_behavior_widgets(condition_combobox, rule_frame, option_entry, index, col, rule_var):
        behavior = condition_combobox.get()
        # <<< FIX 2: Unpack the 6-item tuple (we added the rule_var)
        frame, tag_cb, cond_cb, old_rule_widget, opt_entry, _ = widgets_4[index]
        if old_rule_widget:
            old_rule_widget.destroy()

        if behavior == "транслит":
            rule_widget = ttk.Combobox(rule_frame, textvariable=rule_var, width=18, state="readonly",
                                       values=["uk", "ru", "pl", "hu", "ro"])
            rule_widget.set(rule_var.get() or "uk")
            option_entry.config(state="disabled")
        elif behavior == "числа в слова":
            rule_widget = tk.Frame(rule_frame)
            lang_var = tk.StringVar(value="uk")
            currency_var = tk.StringVar(value="")
            initial_rule = rule_var.get()
            parts = (initial_rule or "").split("|")
            if len(parts) >= 1 and parts[0] in ["uk", "ru", "en"]: lang_var.set(parts[0])
            if len(parts) >= 2 and parts[1] in ["", "UAH", "RUB", "USD", "EUR"]: currency_var.set(parts[1])
            lang_cb = ttk.Combobox(rule_widget, textvariable=lang_var, width=8, state="readonly",
                                   values=["uk", "ru", "en"])
            lang_cb.pack(side="left", padx=(0, 5))
            currency_cb = ttk.Combobox(rule_widget, textvariable=currency_var, width=10,
                                       values=["", "UAH", "RUB", "USD", "EUR"])
            currency_cb.pack(side="left")

            def update_rule_var(*args):
                rule_var.set(f"{lang_var.get()}|{currency_var.get()}")

            lang_var.trace_add("write", update_rule_var)
            currency_var.trace_add("write", update_rule_var)
            update_rule_var()
            option_entry.config(state="disabled")
        else:
            rule_widget = tk.Entry(rule_frame, textvariable=rule_var, width=20)
            if behavior in ["очистить", "CAPS", "верхняя буква", "нижняя буква", "очистить при не выполнении"]:
                rule_widget.config(state="disabled")
            option_entry.config(state="disabled")
            rule_widget.config(state="normal")

        rule_widget.grid(row=0, column=col, padx=5, pady=5)
        # <<< FIX 3: Store the 6-item tuple back into the list
        widgets_4[index] = (frame, tag_cb, cond_cb, rule_widget, opt_entry, rule_var)

    condition_combobox2.bind("<<ComboboxSelected>>",
                             lambda e: toggle_behavior_widgets(condition_combobox2, second_column_frame, option_entry2,
                                                               0, 2, rule_entry_var2))
    toggle_behavior_widgets(condition_combobox2, second_column_frame, option_entry2, 0, 2, rule_entry_var2)

    def add_new_4_item_set():
        base_row = len(widgets_4)
        new_frame = tk.Frame(main_frame)
        new_frame.grid(row=2 + base_row, column=4, columnspan=4, padx=5, pady=5, sticky="w")
        new_tag_var = tk.StringVar()
        new_tag_combobox = ttk.Combobox(new_frame, textvariable=new_tag_var, width=20, state="readonly",
                                        values=sorted_tags)
        new_tag_combobox.grid(row=0, column=0, padx=5, pady=5)
        new_condition_var = tk.StringVar()
        new_condition_combobox = ttk.Combobox(new_frame, textvariable=new_condition_var, width=20, state="readonly",
                                              values=["", "очистить", "очистить при не выполнении", "CAPS",
                                                      "верхняя буква", "нижняя буква", "транслит", "числа в слова",
                                                      "добавить текст в начале", "добавить текст в конце",
                                                      "добавить дней кален", "отнять дней кален", "добавить рабочих дней", "отнять рабочих дней", "обрезать"])
        new_condition_combobox.grid(row=0, column=1, padx=5, pady=5)
        new_rule_var = tk.StringVar()
        new_rule_entry = tk.Entry(new_frame, textvariable=new_rule_var, width=20)  # Placeholder
        new_option_var = tk.StringVar()
        new_option_entry = tk.Entry(new_frame, textvariable=new_option_var, width=20, state="disabled")
        new_option_entry.grid(row=0, column=3, padx=5, pady=5)

        # <<< FIX 4: Store the rule's StringVar in the list for new rows
        widgets_4.append(
            (new_frame, new_tag_combobox, new_condition_combobox, new_rule_entry, new_option_entry, new_rule_var))

        new_condition_combobox.bind("<<ComboboxSelected>>",
                                    lambda e, cb=new_condition_combobox, fr=new_frame, oe=new_option_entry,
                                           idx=base_row, rv=new_rule_var:
                                    toggle_behavior_widgets(cb, fr, oe, idx, 2, rv))
        toggle_behavior_widgets(new_condition_combobox, new_frame, new_option_entry, base_row, 2, new_rule_var)

        buttons_frame_4.grid(row=2 + len(widgets_4), column=4, columnspan=4, padx=5, pady=5, sticky="w")
        new_height = max(230, 230 + max(len(widgets_3), len(widgets_4)) * 40)
        create_rule_window.geometry(f"1070x{new_height}")

    def remove_last_4_item_set():
        if len(widgets_4) > 1:
            # <<< FIX 5: Unpack 6 items when removing a row
            frame_to_destroy, _, _, _, _, _ = widgets_4.pop()
            frame_to_destroy.destroy()
            buttons_frame_4.grid(row=2 + len(widgets_4), column=4, columnspan=4, padx=5, pady=5, sticky="w")
            new_height = max(230, 230 + max(len(widgets_3), len(widgets_4)) * 40)
            create_rule_window.geometry(f"1070x{new_height}")

    buttons_frame_4 = tk.Frame(main_frame)
    buttons_frame_4.grid(row=2 + len(widgets_4), column=4, columnspan=4, padx=5, pady=5, sticky="w")
    ttk.Button(buttons_frame_4, text="Добавить", command=add_new_4_item_set).pack(side="left", padx=(0, 5))
    ttk.Button(buttons_frame_4, text="Удалить", command=remove_last_4_item_set).pack(side="left")

    def validate_and_save():
        rule_name = name_var.get().strip()
        if not rule_name:
            messagebox.showwarning("Ошибка", "Имя правила не может быть пустым.", parent=create_rule_window)
            return

        rules_config = load_json(RULES_CONFIG_PATH, 'rules_config') or []
        if any(rule['name'] == rule_name for rule in rules_config):
            messagebox.showwarning("Ошибка", "Имя правила уже существует.", parent=create_rule_window)
            return

        conditions = []
        for _, tag_cb, cond_cb, _, rule_var in widgets_3:
            tag, condition, rule = tag_cb.get(), cond_cb.get(), rule_var.get()
            if tag and condition:
                if condition not in ["True", "False"] and not rule:
                    messagebox.showwarning("Ошибка", "Введите правило для условия.", parent=create_rule_window)
                    return
                if condition in ["меньше", "больше", "равно"]:
                    try:
                        float(rule)
                    except ValueError:
                        messagebox.showwarning("Ошибка",
                                               "Для условий 'меньше', 'больше' или 'равно' введите только число.",
                                               parent=create_rule_window)
                        return
                conditions.append({'tag': tag, 'condition': condition, 'rule': rule})

        behaviors = []
        # <<< FIX 6: The main fix. Loop over the 6-item tuple and get the rule from the stored StringVar.
        for _, tag_cb, cond_cb, _, _, rule_var in widgets_4:
            tag, condition, rule = tag_cb.get(), cond_cb.get(), rule_var.get()
            if tag and condition:
                if condition not in ["очистить", "CAPS", "верхняя буква", "нижняя буква", "очистить при не выполнении",
                                     "транслит", "числа в слова"] and not rule:
                    messagebox.showwarning("Ошибка", "Введите правило для поведения.", parent=create_rule_window)
                    return
                if condition in ["добавить дней кален", "отнять дней кален"] and not rule.isdigit():
                    messagebox.showwarning("Ошибка",
                                           "Для 'добавить дней кален' или 'отнять дней кален' введите количество дней в виде числа.",
                                           parent=create_rule_window)
                    return
                if condition in ["добавить дней кален", "отнять дней кален", "добавить рабочих дней",
                                 "отнять рабочих дней"] and not rule.isdigit():
                    messagebox.showwarning("Ошибка",
                                           "Для 'добавить дней кален', 'отнять дней кален', 'добавить рабочих дней' или 'отнять рабочих дней' введите количество дней в виде числа.",
                                           parent=create_rule_window)
                    return
                if condition == "обрезать":
                    parts = rule.split(':')
                    if len(parts) != 2 or not (
                            (parts[0] == '' and parts[1].isdigit()) or  # :end format
                            (parts[1] == '' and parts[0].isdigit()) or  # start: format
                            (parts[0].isdigit() and parts[1].isdigit())  # start:end format
                    ):
                        messagebox.showwarning("Ошибка",
                                               "Для 'обрезать' введите диапазон в формате 'start:end', ':end' или 'start:' с числами.",
                                               parent=create_rule_window)
                        return
                behaviors.append({'tag': tag, 'condition': condition, 'rule': rule})

        if not conditions and not behaviors:
            messagebox.showwarning("Ошибка", "Правило должно содержать хотя бы одно условие или поведение.",
                                   parent=create_rule_window)
            return

        new_rule = {'name': rule_name, 'conditions': conditions, 'behaviors': behaviors}
        rules_config.append(new_rule)
        save_json(RULES_CONFIG_PATH, rules_config)
        messagebox.showinfo("Успех", f"Правило '{rule_name}' успешно создано.", parent=create_rule_window)
        update_rules_listbox(rules_config, listbox)
        create_rule_window.destroy()

    button_frame = tk.Frame(create_rule_window)
    button_frame.pack(pady=20, fill=tk.X)
    button_frame.columnconfigure(0, weight=1)
    button_frame.columnconfigure(1, weight=1)
    tk.Button(button_frame, text="ОК", width=10, command=validate_and_save).grid(row=0, column=0, padx=5)
    tk.Button(button_frame, text="ОТМЕНА", width=10, command=create_rule_window.destroy).grid(row=0, column=1, padx=5)


def open_edit_rule_window(listbox, constructor_window):
    selected_item = listbox.selection()
    if not selected_item:
        messagebox.showwarning("Ошибка", "Выберите правило для редактирования", parent=constructor_window)
        return

    item_data = listbox.item(selected_item)
    rule_name = item_data['values'][0]

    rules_config = load_json(RULES_CONFIG_PATH, 'rules_config')
    rule_to_edit = next((rule for rule in rules_config if rule['name'] == rule_name), None)

    if not rule_to_edit:
        messagebox.showerror("Ошибка", f"Правило '{rule_name}' не найдено.")
        return

    edit_rule_window = tk.Toplevel(constructor_window)
    edit_rule_window.title(f"Изменить правило: {rule_name}")
    edit_rule_window.geometry("1070x230")
    edit_rule_window.resizable(False, False)
    edit_rule_window.focus_set()
    edit_rule_window.grab_set()

    edit_rule_window._vars_to_keep = []
    sorted_tags = sorted(list(set(get_all_tags_for_constructor())))
    sorted_tags.insert(0, '')

    main_frame = tk.Frame(edit_rule_window)
    main_frame.pack(padx=10, pady=5, fill=tk.BOTH, expand=True)

    widgets_3 = []
    widgets_4 = []

    tk.Label(main_frame, text="Имя правила:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
    name_var = tk.StringVar(value=rule_name)
    tk.Entry(main_frame, textvariable=name_var, width=30).grid(row=0, column=1, columnspan=5, padx=5, pady=5,
                                                               sticky="w")

    tk.Label(main_frame, text="Теги", anchor="center").grid(row=1, column=0, padx=5, pady=5)
    tk.Label(main_frame, text="Условие", anchor="center").grid(row=1, column=1, padx=5, pady=5)
    tk.Label(main_frame, text="Правило", anchor="center").grid(row=1, column=2, padx=5, pady=5)

    buttons_frame_3 = tk.Frame(main_frame)
    buttons_frame_4 = tk.Frame(main_frame)

    def toggle_rule_entry_state(condition_combobox, rule_entry):
        selected_condition = condition_combobox.get()
        rule_entry.config(state="disabled" if selected_condition in ["True", "False"] else "normal")

    def add_new_3_item_set(pre_populated_data=None):
        base_row = len(widgets_3)
        new_frame = tk.Frame(main_frame)
        new_frame.grid(row=2 + base_row, column=0, columnspan=3, padx=5, pady=5, sticky="w")
        new_tag_var = tk.StringVar()
        new_tag_combobox = ttk.Combobox(new_frame, textvariable=new_tag_var, width=20, state="readonly",
                                        values=sorted_tags)
        new_tag_combobox.grid(row=0, column=0, padx=5, pady=5)
        new_condition_var = tk.StringVar()
        new_condition_combobox = ttk.Combobox(new_frame, textvariable=new_condition_var, width=20, state="readonly",
                                              values=["", "содержит", "начинается с", "заканчивается на", "меньше",
                                                      "больше", "равно", "True", "False"])
        new_condition_combobox.grid(row=0, column=1, padx=5, pady=5)
        new_rule_var = tk.StringVar()
        new_rule_entry = tk.Entry(new_frame, textvariable=new_rule_var, width=20)
        new_rule_entry.grid(row=0, column=2, padx=5, pady=5)
        widgets_3.append((new_frame, new_tag_combobox, new_condition_combobox, new_rule_entry, new_rule_var))
        edit_rule_window._vars_to_keep.extend([new_tag_var, new_condition_var, new_rule_var])
        new_condition_combobox.bind("<<ComboboxSelected>>",
                                    lambda e: toggle_rule_entry_state(new_condition_combobox, new_rule_entry))
        if pre_populated_data:
            new_tag_var.set(pre_populated_data.get('tag', ''))
            new_condition_var.set(pre_populated_data.get('condition', ''))
            new_rule_var.set(pre_populated_data.get('rule', ''))
            toggle_rule_entry_state(new_condition_combobox, new_rule_entry)
        buttons_frame_3.grid(row=2 + len(widgets_3), column=0, columnspan=3, padx=5, pady=5, sticky="w")
        new_height = max(230, 230 + max(len(widgets_3), len(widgets_4)) * 40)
        edit_rule_window.geometry(f"1070x{new_height}")

    def remove_last_3_item_set():
        if len(widgets_3) > 1:
            widgets_3.pop()[0].destroy()
            buttons_frame_3.grid(row=2 + len(widgets_3), column=0, columnspan=3, padx=5, pady=5, sticky="w")
            new_height = max(230, 230 + max(len(widgets_3), len(widgets_4)) * 40)
            edit_rule_window.geometry(f"1070x{new_height}")

    if rule_to_edit.get('conditions'):
        for condition in rule_to_edit.get('conditions'):
            add_new_3_item_set(condition)
    else:
        add_new_3_item_set()
    ttk.Button(buttons_frame_3, text="Добавить", command=lambda: add_new_3_item_set()).pack(side="left", padx=(0, 5))
    ttk.Button(buttons_frame_3, text="Удалить", command=remove_last_3_item_set).pack(side="left")

    ttk.Separator(main_frame, orient='vertical').grid(row=1, column=3, rowspan=100, sticky="ns", padx=10)

    tk.Label(main_frame, text="Теги", anchor="center").grid(row=1, column=4, padx=5, pady=5)
    tk.Label(main_frame, text="Условие", anchor="center").grid(row=1, column=5, padx=5, pady=5)
    tk.Label(main_frame, text="Правило", anchor="center").grid(row=1, column=6, padx=5, pady=5)
    tk.Label(main_frame, text="Опция", anchor="center").grid(row=1, column=7, padx=5, pady=5)

    def toggle_behavior_widgets(condition_combobox, rule_frame, option_entry, index, col, rule_var):
        behavior = condition_combobox.get()
        frame, tag_cb, cond_cb, old_rule_widget, opt_entry, _ = widgets_4[index]
        if old_rule_widget:
            old_rule_widget.destroy()

        if behavior == "транслит":
            rule_widget = ttk.Combobox(rule_frame, textvariable=rule_var, width=18, state="readonly")
            rule_widget['values'] = ["uk", "ru", "pl", "hu", "ro"]
            rule_widget.set(rule_var.get() or "uk")
            option_entry.config(state="disabled")
        elif behavior == "числа в слова":
            rule_widget = tk.Frame(rule_frame)
            lang_var = tk.StringVar(value="uk")
            currency_var = tk.StringVar(value="")
            initial_rule = rule_var.get()
            parts = (initial_rule or "").split("|")
            if len(parts) >= 1 and parts[0] in ["uk", "ru", "en"]:
                lang_var.set(parts[0])
            if len(parts) >= 2 and parts[1] in ["", "UAH", "RUB", "USD", "EUR"]:
                currency_var.set(parts[1])
            lang_cb = ttk.Combobox(rule_widget, textvariable=lang_var, width=8, state="readonly",
                                   values=["uk", "ru", "en"])
            lang_cb.pack(side="left", padx=(0, 5))
            currency_cb = ttk.Combobox(rule_widget, textvariable=currency_var, width=10,
                                       values=["", "UAH", "RUB", "USD", "EUR"])
            currency_cb.pack(side="left")

            def update_rule_var(*args):
                rule_var.set(f"{lang_var.get()}|{currency_var.get()}")

            lang_var.trace_add("write", update_rule_var)
            currency_var.trace_add("write", update_rule_var)
            update_rule_var()
            option_entry.config(state="disabled")
        else:
            rule_widget = tk.Entry(rule_frame, textvariable=rule_var, width=20)
            if behavior in ["очистить", "CAPS", "верхняя буква", "нижняя буква", "очистить при не выполнении"]:
                rule_widget.config(state="disabled")
                option_entry.config(state="disabled")
            else:
                rule_widget.config(state="normal")
                option_entry.config(state="disabled")

        rule_widget.grid(row=0, column=col, padx=5, pady=5)
        widgets_4[index] = (frame, tag_cb, cond_cb, rule_widget, opt_entry, rule_var)

    def add_new_4_item_set(pre_populated_data=None):
        base_row = len(widgets_4)
        new_frame = tk.Frame(main_frame)
        new_frame.grid(row=2 + base_row, column=4, columnspan=4, padx=5, pady=5, sticky="w")
        new_tag_var = tk.StringVar()
        new_tag_combobox = ttk.Combobox(new_frame, textvariable=new_tag_var, width=20, state="readonly",
                                        values=sorted_tags)
        new_tag_combobox.grid(row=0, column=0, padx=5, pady=5)
        new_condition_var = tk.StringVar()
        new_condition_combobox = ttk.Combobox(new_frame, textvariable=new_condition_var, width=20, state="readonly")
        new_condition_combobox['values'] = [
            "", "очистить", "очистить при не выполнении", "CAPS", "верхняя буква", "нижняя буква", "транслит",
            "числа в слова", "добавить текст в начале", "добавить текст в конце", "добавить дней кален", "отнять дней кален",
            "добавить рабочих дней", "отнять рабочих дней", "обрезать"
        ]
        new_condition_combobox.grid(row=0, column=1, padx=5, pady=5)
        new_rule_var = tk.StringVar()
        new_rule_entry = tk.Entry(new_frame, textvariable=new_rule_var, width=20)  # Placeholder
        new_option_var = tk.StringVar()
        new_option_entry = tk.Entry(new_frame, textvariable=new_option_var, width=20, state="disabled")
        new_option_entry.grid(row=0, column=3, padx=5, pady=5)

        widgets_4.append(
            (new_frame, new_tag_combobox, new_condition_combobox, new_rule_entry, new_option_entry, new_rule_var))
        edit_rule_window._vars_to_keep.extend([new_tag_var, new_condition_var, new_rule_var, new_option_var])

        new_condition_combobox.bind("<<ComboboxSelected>>",
                                    lambda e, cb=new_condition_combobox, fr=new_frame, oe=new_option_entry,
                                           idx=base_row, rv=new_rule_var:
                                    toggle_behavior_widgets(cb, fr, oe, idx, 2, rv))
        if pre_populated_data:
            new_tag_var.set(pre_populated_data.get('tag', ''))
            new_condition_var.set(pre_populated_data.get('condition', ''))
            new_rule_var.set(pre_populated_data.get('rule', ''))
            new_option_var.set(pre_populated_data.get('option', ''))

        toggle_behavior_widgets(new_condition_combobox, new_frame, new_option_entry, base_row, 2, new_rule_var)

        buttons_frame_4.grid(row=2 + len(widgets_4), column=4, columnspan=4, padx=5, pady=5, sticky="w")
        new_height = max(230, 230 + max(len(widgets_3), len(widgets_4)) * 40)
        edit_rule_window.geometry(f"1070x{new_height}")

    def remove_last_4_item_set():
        if len(widgets_4) > 1:
            frame_to_destroy, _, _, _, _, _ = widgets_4.pop()
            frame_to_destroy.destroy()
            buttons_frame_4.grid(row=2 + len(widgets_4), column=4, columnspan=4, padx=5, pady=5, sticky="w")
            new_height = max(230, 230 + max(len(widgets_3), len(widgets_4)) * 40)
            edit_rule_window.geometry(f"1070x{new_height}")

    if rule_to_edit.get('behaviors'):
        for behavior in rule_to_edit.get('behaviors'):
            add_new_4_item_set(behavior)
    else:
        add_new_4_item_set()

    ttk.Button(buttons_frame_4, text="Добавить", command=lambda: add_new_4_item_set()).pack(side="left", padx=(0, 5))
    ttk.Button(buttons_frame_4, text="Удалить", command=remove_last_4_item_set).pack(side="left")

    def validate_and_save_edit():
        new_rule_name = name_var.get().strip()
        if not new_rule_name:
            messagebox.showwarning("Ошибка", "Имя правила не может быть пустым.", parent=edit_rule_window)
            return
        rules_config = load_json(RULES_CONFIG_PATH, 'rules_config') or []
        if new_rule_name != rule_name and any(rule['name'] == new_rule_name for rule in rules_config):
            messagebox.showwarning("Ошибка", "Имя правила уже существует.", parent=edit_rule_window)
            return

        conditions = []
        for _, tag_cb, cond_cb, _, rule_var in widgets_3:
            tag, condition, rule = tag_cb.get(), cond_cb.get(), rule_var.get()
            if tag and condition:
                if condition not in ["True", "False"] and not rule:
                    messagebox.showwarning("Ошибка", "Введите правило для условия.", parent=edit_rule_window)
                    return
                if condition in ["меньше", "больше", "равно"]:
                    try:
                        float(rule)
                    except ValueError:
                        messagebox.showwarning("Ошибка",
                                               "Для условий 'меньше', 'больше' или 'равно' введите только число.",
                                               parent=edit_rule_window)
                        return
                conditions.append({'tag': tag, 'condition': condition, 'rule': rule})

        behaviors = []
        for _, tag_cb, cond_cb, _, _, rule_var in widgets_4:
            tag, condition, rule = tag_cb.get(), cond_cb.get(), rule_var.get()
            if tag and condition:
                if condition not in ["очистить", "CAPS", "верхняя буква", "нижняя буква",
                                     "очистить при не выполнении", "транслит", "числа в слова"] and not rule:
                    messagebox.showwarning("Ошибка", "Введите правило для поведения.", parent=edit_rule_window)
                    return
                if condition in ["добавить дней кален", "отнять дней кален"] and not rule.isdigit():
                    messagebox.showwarning("Ошибка",
                                           "Для 'добавить дней кален' или 'отнять дней кален' введите количество дней в виде числа.",
                                           parent=edit_rule_window)
                    return
                if condition in ["добавить дней кален", "отнять дней кален", "добавить рабочих дней",
                                 "отнять рабочих дней"] and not rule.isdigit():
                    messagebox.showwarning("Ошибка",
                                           "Для 'добавить дней кален', 'отнять дней кален', 'добавить рабочих дней' или 'отнять рабочих дней' введите количество дней в виде числа.",
                                           parent=edit_rule_window)
                    return
                if condition == "обрезать":
                    parts = rule.split(':')
                    if len(parts) != 2 or not (
                            (parts[0] == '' and parts[1].isdigit()) or  # :end format
                            (parts[1] == '' and parts[0].isdigit()) or  # start: format
                            (parts[0].isdigit() and parts[1].isdigit())  # start:end format
                    ):
                        messagebox.showwarning("Ошибка",
                                               "Для 'обрезать' введите диапазон в формате 'start:end', ':end' или 'start:' с числами.",
                                               parent=edit_rule_window)
                        return
                behaviors.append({'tag': tag, 'condition': condition, 'rule': rule})

        updated_rules = [rule for rule in rules_config if rule['name'] != rule_name]
        updated_rules.append({'name': new_rule_name, 'conditions': conditions, 'behaviors': behaviors})
        save_json(RULES_CONFIG_PATH, updated_rules)
        messagebox.showinfo("Успех", f"Правило '{new_rule_name}' успешно обновлено.", parent=edit_rule_window)
        update_rules_listbox(updated_rules, listbox)
        edit_rule_window.destroy()

    button_frame = tk.Frame(edit_rule_window)
    button_frame.pack(pady=20, fill=tk.X, side=tk.BOTTOM)
    button_frame.columnconfigure(0, weight=1)
    button_frame.columnconfigure(1, weight=1)
    tk.Button(button_frame, text="ОК", width=10, command=validate_and_save_edit).grid(row=0, column=0, padx=5)
    tk.Button(button_frame, text="ОТМЕНА", width=10, command=edit_rule_window.destroy).grid(row=0, column=1, padx=5)


def delete_rule(listbox, parent_window, rules_file=None):
    if rules_file is None:
        rules_file = RULES_CONFIG_PATH
    """Deletes one or more selected rules."""
    selected_items = listbox.selection()
    if not selected_items:
        messagebox.showwarning("Ошибка", "Выберите правило(а) для удаления", parent=parent_window)
        return

    # Collect the names of all selected rules
    rules_to_delete_names = [listbox.item(item_id)['values'][0] for item_id in selected_items]

    # Format a user-friendly confirmation message
    names_str = "\n- ".join(rules_to_delete_names)
    response = messagebox.askyesno("Подтверждение",
                                   f"Вы уверены, что хотите удалить следующие правила?\n\n- {names_str}",
                                   parent=parent_window)
    if not response:
        return

    try:
        existing_rules = load_json(rules_file, 'rules_config')
        if not isinstance(existing_rules, list):  # Basic validation
            existing_rules = []

        # Use a set for efficient filtering
        rules_to_delete_set = set(rules_to_delete_names)
        updated_rules = [rule for rule in existing_rules if rule.get('name') not in rules_to_delete_set]

        save_json(rules_file, updated_rules)

        # Update the listbox with the new data
        update_rules_listbox(updated_rules, listbox)
        messagebox.showinfo("Успех", "Выбранные правила были успешно удалены.", parent=parent_window)

    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось сохранить изменения: {str(e)}", parent=parent_window)


# --- Startup logic ---

# Determine the base directory
if getattr(sys, 'frozen', False):
    BASE_DIR = sys._MEIPASS
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

ICON_DIR = os.path.join(BASE_DIR, 'icon')
ICON_PATH = os.path.join(ICON_DIR, "bee.ico")


def start_main_window():
    """Builds and launches the main application window after a project is selected."""
    global window, dynamic_frame

    # Create a single Tkinter window instance
    window = tk.Tk()
    window.title("Doxy v2.1")
    # 1. CHANGE: Allow resizing so scrollbars can be useful
    window.resizable(True, True)
    window.protocol("WM_DELETE_WINDOW", on_closing)

    # Set main window icon
    try:
        window.iconbitmap(ICON_PATH)
    except tk.TclError:
        print("Warning: Icon file not found or invalid format.")

    # Apply icon automatically to all Toplevel windows
    _original_toplevel_init = tk.Toplevel.__init__

    def _custom_toplevel_init(self, *args, **kwargs):
        _original_toplevel_init(self, *args, **kwargs)
        try:
            self.iconbitmap(ICON_PATH)
        except tk.TclError:
            pass

    tk.Toplevel.__init__ = _custom_toplevel_init

    # --- Main layout frames (REPLACED WITH SCROLLABLE CONTENT) ---

    # Container frame to hold the canvas and scrollbars
    content_container = tk.Frame(window)
    content_container.pack(fill="both", expand=True)

    # Vertical Scrollbar
    v_scroll = ttk.Scrollbar(content_container, orient="vertical")
    v_scroll.pack(side="right", fill="y")

    # Horizontal Scrollbar
    h_scroll = ttk.Scrollbar(content_container, orient="horizontal")
    h_scroll.pack(side="bottom", fill="x")

    # Canvas (The viewport)
    canvas = tk.Canvas(content_container, yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)
    canvas.pack(side="left", fill="both", expand=True)

    # Link scrollbars to canvas
    v_scroll.config(command=canvas.yview)
    h_scroll.config(command=canvas.xview)

    # Create the dynamic_frame INSIDE the canvas
    dynamic_frame = tk.Frame(canvas, padx=10, pady=10)
    canvas.create_window((0, 0), window=dynamic_frame, anchor="nw")

    # Critical: Update scroll region whenever dynamic_frame changes size (i.e., tags are added)
    def on_frame_configure(event):
        canvas.configure(scrollregion=canvas.bbox("all"))

    dynamic_frame.bind("<Configure>", on_frame_configure)

    # Optional: Enable mouse wheel scrolling
    if sys.platform.startswith('win') or sys.platform.startswith('darwin'):
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1 * (e.delta / 120)), "units"))
    elif sys.platform.startswith('linux'):
        canvas.bind_all("<Button-4>", lambda e: canvas.yview_scroll(-1, "units"))
        canvas.bind_all("<Button-5>", lambda e: canvas.yview_scroll(1, "units"))

    ttk.Separator(window, orient='horizontal').pack(fill='x', pady=5)

    bottom_frame = tk.Frame(window)
    # 2. CHANGE: Anchor the bottom frame to the bottom
    bottom_frame.pack(fill="x", padx=10, pady=(0, 10), side=BOTTOM)

    # --- Bottom Buttons ---
    constructor_button = ttk.Button(bottom_frame, text="Конструктор", command=open_constructor_window)
    report_button = ttk.Button(bottom_frame, text="Сформировать", command=submit_and_save)

    def switch_to_projects():
        """
        Показывает диалог сохранения/стирания данных перед переключением на окно
        выбора проектов.
        """
        answer = messagebox.askyesnocancel(
            "Переход к проектам",
            "Вы хотите сохранить введённые данные текущего проекта?\n"
            "Да = сохранить данные\n"
            "Нет = очистить все данные\n"
            "Отмена = вернуться"
        )
        if answer is None:  # Отмена
            return
        elif answer:  # Да -> Сохранить и перейти
            save_input_state()
            window.destroy()
            show_welcome_window()
        else:  # Нет -> Очистить и перейти
            clear_all_inputs()
            save_json(INPUT_STATE_PATH, {})
            window.destroy()
            show_welcome_window()

    projects_button = ttk.Button(bottom_frame, text="Проекты", command=switch_to_projects)

    # pack buttons
    ttk.Button(bottom_frame, text="Импорт", command=import_fields).pack(side=LEFT, padx=5, pady=5, fill=X, expand=True)
    constructor_button.pack(side=LEFT, padx=5, pady=5, fill=X, expand=True)
    report_button.pack(side=LEFT, padx=5, pady=5, fill=X, expand=True)
    projects_button.pack(side=LEFT, padx=5, pady=5, fill=X, expand=True)

    # --- Initial Load ---
    window.bind_all("<Key>", _onKeyRelease, "+")
    load_all_dynamic_widgets()

    window.mainloop()

def show_welcome_window():
    root = tk.Tk()
    root.resizable(False, False)
    root.title("Выбор проекта")
    root.geometry("400x300")

    try:
        root.iconbitmap(ICON_PATH)
    except tk.TclError:
        print("Warning: Icon file not found or invalid format.")

    # Apply icon automatically to all Toplevel windows
    _original_toplevel_init = tk.Toplevel.__init__
    def _custom_toplevel_init(self, *args, **kwargs):
        _original_toplevel_init(self, *args, **kwargs)
        try:
            self.iconbitmap(ICON_PATH)
        except tk.TclError:
            pass
    tk.Toplevel.__init__ = _custom_toplevel_init

    def refresh_list():
        for i in project_listbox.get_children():
            project_listbox.delete(i)
        for p in load_projects()["projects"]:
            project_listbox.insert(
                "", "end",
                values=(p["name"], "Да" if p["autoload"] else "Нет")
            )

    def on_create():
        name = askstring_localized(".", "Введите имя проекта:")
        if name:
            if create_project(name):
                set_current_project(name)
                projects = load_projects()
                projects["last_opened"] = name
                save_projects(projects)
                root.destroy()
                start_main_window()
    def on_load():
        sel = project_listbox.selection()
        if not sel:
            return
        name = project_listbox.item(sel[0])["values"][0]
        set_current_project(name)
        projects = load_projects()
        projects["last_opened"] = name
        save_projects(projects)
        root.destroy()
        start_main_window()   # <-- Launch app after project chosen

    def on_delete():
        sel = project_listbox.selection()
        if not sel:
            return
        name = project_listbox.item(sel[0])["values"][0]
        if messagebox.askyesno("Подтверждение", f"Удалить проект '{name}'?"):
            delete_project(name)
            refresh_list()

    def on_autoload():
        sel = project_listbox.selection()
        if not sel:
            return
        name = project_listbox.item(sel[0])["values"][0]

        projects = load_projects()
        for p in projects["projects"]:
            if p["name"] == name:
                # toggle
                new_state = not p.get("autoload", False)
                p["autoload"] = new_state
            else:
                # all other projects must be off
                p["autoload"] = False
        save_projects(projects)
        refresh_list()

    project_listbox = ttk.Treeview(
        root,
        columns=("Имя", "Загружать при старте"),
        show="headings"
    )
    project_listbox.heading("Имя", text="Имя")
    project_listbox.heading("Загружать при старте", text="Загружать при старте")
    project_listbox.column("Имя", anchor="center", width=200)
    project_listbox.column("Загружать при старте", anchor="center", width=200)
    project_listbox.pack(fill="both", expand=True)

    btn_frame = tk.Frame(root)
    btn_frame.pack(pady=10)

    tk.Button(btn_frame, text="Создать", command=on_create).grid(row=0, column=0, padx=5)
    tk.Button(btn_frame, text="Загрузить", command=on_load).grid(row=0, column=1, padx=5)
    tk.Button(btn_frame, text="Удалить", command=on_delete).grid(row=0, column=2, padx=5)
    tk.Button(btn_frame, text="Автозагрузка", command=on_autoload).grid(row=0, column=3, padx=5)

    refresh_list()
    root.mainloop()


# --- Startup ---
if __name__ == "__main__":
    projects = load_projects()
    autoload_project = get_autoload_project()

    if autoload_project:
        set_current_project(autoload_project)
        start_main_window()
    else:
        show_welcome_window()
